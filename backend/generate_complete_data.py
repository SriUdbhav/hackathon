#!/usr/bin/env python3
"""
Generate Complete Realistic Demo Dataset for EduStudent Sight
- 100 Students (covering Low, Moderate, High, and Critical Risk cohorts)
- 50 Faculty & Mentors (30 Faculty with specialized subjects + 20 Mentors)
- 10 Admins (Dean, HODs, Exam Incharges, Welfare Directors, System Admins)
- 500 Subject Marks (5 core semester subjects per student with granular marks)
- Extracurricular Activities & Student Participation Records
- 30+ Interventions (Completed, In Progress, Completion Requested, Revision Needed)
- Signup Requests & Notifications
- Exports: .CSV files & Multi-sheet Excel workbook in data/
- Direct SQLite Database initialization & seeding
"""

import os
import sys
import random
import sqlite3
import csv
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Ensure deterministic yet realistic data
random.seed(42)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(BASE_DIR), "data")
DB_PATH = os.path.join(BASE_DIR, "database.db")
os.makedirs(DATA_DIR, exist_ok=True)

# -------------------------------------------------------------
# 1. NAME POOLS & REGIONAL DISTRIBUTIONS
# -------------------------------------------------------------

FIRST_NAMES_MALE = [
    "Sri Udbhav", "Hemanth", "Gopi", "Arjun", "Aditya", "Rohan", "Vikram", "Siddharth",
    "Sai Krishna", "Praveen", "Karthik", "Varun", "Nikhil", "Rahul", "Tanmay", "Abhishek",
    "Anish", "Manish", "Tarun", "Harsh", "Pranav", "Deepak", "Chaitanya", "Vamsi", "Ravi",
    "Surya", "Teja", "Kiran", "Naveen", "Bhargav", "Akhil", "Charan", "Lokesh", "Suresh",
    "Rajesh", "Ganesh", "Mahesh", "Dinesh", "Kalyan", "Shravan", "Vinay", "Suhas", "Anil",
    "Sunil", "Pavan", "Vikas", "Gautam", "Avinash", "Rohit", "Ashwin", "Dev", "Yash"
]

FIRST_NAMES_FEMALE = [
    "Sneha", "Ananya", "Pooja", "Divya", "Swathi", "Kavya", "Keerthi", "Harini",
    "Sravani", "Bhavya", "Rhea", "Isha", "Meera", "Deepika", "Shreya", "Nandini",
    "Tanvi", "Sanjana", "Mounika", "Tejaswi", "Preethi", "Aishwarya", "Lahari",
    "Navya", "Sahithi", "Pavani", "Gowri", "Sushma", "Geetha", "Rashmi", "Varsha",
    "Gayathri", "Pallavi", "Lavanya", "Sirisha", "Sandhya", "Radhika", "Madhuri",
    "Rupa", "Alekhya", "Sindhu", "Harika", "Aparna", "Archana", "Manasa", "Jyothi",
    "Ramya", "Supriya", "Vyshnavi", "Pranathi"
]

LAST_NAMES = [
    "Vangapandu", "Yalamanchili", "Tadepalli", "Patel", "Sharma", "Reddy", "Rao",
    "Verma", "Gupta", "Nair", "Iyer", "Choudhury", "Menon", "Joshi", "Bhat", "Kulkarni",
    "Deshmukh", "Chowdary", "Kolisetty", "Guduru", "Mikkilineni", "Kondapalli", "Atluri",
    "Garlapati", "Yarram", "Penumaka", "Mullapudi", "Gottipati", "Daggubati", "Rayapati",
    "Chalasani", "Garikapati", "Bandaru", "Nalluri", "Uppalapati", "Kandula", "Vemuri"
]

PLACES_REGIONS = [
    ("Hyderabad", "South India", "Telugu"),
    ("Vijayawada", "South India", "Telugu"),
    ("Guntur", "South India", "Telugu"),
    ("Visakhapatnam", "South India", "Telugu"),
    ("Tirupati", "South India", "Telugu"),
    ("Bengaluru", "South India", "Kannada"),
    ("Chennai", "South India", "Tamil"),
    ("Kochi", "South India", "Malayalam"),
    ("Mumbai", "West India", "Marathi"),
    ("Pune", "West India", "Marathi"),
    ("Ahmedabad", "West India", "Gujarati"),
    ("New Delhi", "North India", "Hindi"),
    ("Jaipur", "North India", "Hindi"),
    ("Lucknow", "North India", "Hindi"),
    ("Kolkata", "East India", "Bengali"),
    ("Bhubaneswar", "East India", "Odia"),
    ("Patna", "East India", "Hindi")
]

COURSES = ["CSE", "CSE (AI & ML)", "CSE (Data Science)", "CSE (Cyber Security)", "IT", "ECE"]
YEARS = ["2nd Year", "3rd Year", "4th Year"]

SUBJECT_POOL = [
    {"code": "CS201", "name": "Database Management Systems", "short_name": "DBMS", "credits": 4, "year": "2nd Year", "sem": 1},
    {"code": "CS202", "name": "Operating Systems", "short_name": "OS", "credits": 4, "year": "2nd Year", "sem": 1},
    {"code": "CS203", "name": "Computer Networks", "short_name": "CN", "credits": 4, "year": "2nd Year", "sem": 1},
    {"code": "CS204", "name": "Software Engineering & Agile", "short_name": "SE", "credits": 3, "year": "2nd Year", "sem": 1},
    {"code": "MA201", "name": "Discrete Mathematics & Stats", "short_name": "Math-III", "credits": 4, "year": "2nd Year", "sem": 1},
    {"code": "CS301", "name": "Design & Analysis of Algorithms", "short_name": "DAA", "credits": 4, "year": "3rd Year", "sem": 1},
    {"code": "CS302", "name": "Artificial Intelligence & ML", "short_name": "AI/ML", "credits": 4, "year": "3rd Year", "sem": 1},
    {"code": "CS303", "name": "Web Technologies & Cloud", "short_name": "Web/Cloud", "credits": 3, "year": "3rd Year", "sem": 1},
]

EXTRACURRICULARS_POOL = [
    ("HACK01", "Inter-College Hackathon 2026", "Hackathon", "48-hour competitive innovation hackathon"),
    ("ROBO01", "Autonomous Robotics Society", "Robotics", "Designing hardware bots & IoT sensors"),
    ("AI_SOC", "AI & Machine Learning Club", "Tech Club", "Generative AI workshops, LLM research, kaggle contests"),
    ("CODE01", "Competitive Coding Guild", "Technical", "Algorithm sprint competitions and Codeforces rankings"),
    ("GDSC01", "Google Developer Student Club", "Technical", "Cloud architecture, Flutter, Android, Web platforms"),
    ("CULT01", "Vignan Annual Cultural Fest", "Cultural", "Music, drama, fine arts, stage coordination"),
    ("SPOR01", "Inter-Department Athletics", "Sports", "Cricket, football, badminton, basketball championships"),
    ("NSS01",  "National Service Scheme", "Volunteering", "Rural community empowerment & health camps"),
    ("ECELL",  "Student Entrepreneurship Cell", "Leadership", "Startups pitching, incubation & venture building"),
    ("IEEE01", "IEEE Student Branch Chapter", "Academic", "Publishing research papers, IEEE symposiums")
]


# -------------------------------------------------------------
# 2. GENERATE 100 STUDENTS
# -------------------------------------------------------------

def generate_students(n=100):
    students = []
    
    # Pre-seed iconic 5 students for guaranteed demo login consistency
    students.append({
        "id": "25CS001",
        "name": "V. Sri Udbhav",
        "gender": "Male",
        "course": "CSE",
        "year": "2nd Year",
        "cgpa": 8.4,
        "credits": 24,
        "attendance": 84,
        "lms_score": 88,
        "father": "Ramesh Kumar V",
        "mother": "Lakshmi V",
        "mother_tongue": "Telugu",
        "place": "Hyderabad",
        "region": "South India",
        "country": "India",
        "email": "sriudbhav.25cs@vignan.ac.in",
        "phone": "+91 98480 12345"
    })
    students.append({
        "id": "25CS002",
        "name": "Y. Hemanth Reddy",
        "gender": "Male",
        "course": "CSE",
        "year": "2nd Year",
        "cgpa": 7.4,
        "credits": 23,
        "attendance": 68,
        "lms_score": 60,
        "father": "Reddy Kumar Y",
        "mother": "Padma Y",
        "mother_tongue": "Telugu",
        "place": "Vijayawada",
        "region": "South India",
        "country": "India",
        "email": "hemanth.25cs@vignan.ac.in",
        "phone": "+91 98480 23456"
    })
    students.append({
        "id": "25CS003",
        "name": "T. Gopi",
        "gender": "Male",
        "course": "CSE",
        "year": "2nd Year",
        "cgpa": 7.8,
        "credits": 22,
        "attendance": 74,
        "lms_score": 70,
        "father": "Srinivas Rao T",
        "mother": "Anitha T",
        "mother_tongue": "Telugu",
        "place": "Guntur",
        "region": "South India",
        "country": "India",
        "email": "gopi.25cs@vignan.ac.in",
        "phone": "+91 98480 34567"
    })
    students.append({
        "id": "25CS004",
        "name": "Sneha Rao",
        "gender": "Female",
        "course": "CSE",
        "year": "2nd Year",
        "cgpa": 8.9,
        "credits": 25,
        "attendance": 92,
        "lms_score": 95,
        "father": "Rao Kumar",
        "mother": "Sunitha Rao",
        "mother_tongue": "Telugu",
        "place": "Hyderabad",
        "region": "South India",
        "country": "India",
        "email": "sneha.25cs@vignan.ac.in",
        "phone": "+91 98480 45678"
    })
    students.append({
        "id": "25CS005",
        "name": "Arjun Patel",
        "gender": "Male",
        "course": "CSE",
        "year": "2nd Year",
        "cgpa": 6.8,
        "credits": 20,
        "attendance": 61,
        "lms_score": 50,
        "father": "Mahesh Patel",
        "mother": "Kavitha Patel",
        "mother_tongue": "Hindi",
        "place": "Mumbai",
        "region": "West India",
        "country": "India",
        "email": "arjun.25cs@vignan.ac.in",
        "phone": "+91 98480 56789"
    })

    # Generate remaining 95 students across planned risk distribution
    # Profile buckets:
    # - 50 Low Risk (Safe): Attd 80-98%, CGPA 8.0-9.8, LMS 75-98% -> Risk 5-30%
    # - 25 Moderate Risk (Warning): Attd 65-79%, CGPA 6.5-7.8, LMS 55-72% -> Risk 35-58%
    # - 20 Critical/High Risk: Attd 28-56%, CGPA 3.5-5.6, LMS 18-45% -> Risk 60-88%
    
    risk_plan = (
        [("low", 1)] * 50 +
        [("med", 2)] * 25 +
        [("high", 3)] * 20
    )
    random.shuffle(risk_plan)

    for i, (risk_tier, _) in enumerate(risk_plan, start=6):
        sid = f"25CS{i:03d}"
        gender = "Male" if random.random() < 0.52 else "Female"
        first = random.choice(FIRST_NAMES_MALE if gender == "Male" else FIRST_NAMES_FEMALE)
        last = random.choice(LAST_NAMES)
        full_name = f"{first} {last}"
        
        place, region, mother_tongue = random.choice(PLACES_REGIONS)
        father_name = f"{random.choice(FIRST_NAMES_MALE)} {last}"
        mother_name = f"{random.choice(FIRST_NAMES_FEMALE)} {last}"
        course = random.choice(COURSES)
        year = random.choice(YEARS)
        
        if risk_tier == "low":
            attd = random.randint(80, 98)
            cgpa = round(random.uniform(8.0, 9.7), 2)
            lms = random.randint(75, 96)
            credits_val = random.randint(23, 27)
        elif risk_tier == "med":
            attd = random.randint(65, 79)
            cgpa = round(random.uniform(6.5, 7.8), 2)
            lms = random.randint(55, 72)
            credits_val = random.randint(20, 24)
        else: # high / critical
            attd = random.randint(28, 56)
            cgpa = round(random.uniform(3.5, 5.6), 2)
            lms = random.randint(18, 45)
            credits_val = random.randint(14, 19)

        slug = f"{first.lower().replace(' ', '')}.{sid.lower()}@vignan.ac.in"
        phone = f"+91 98480 {random.randint(10000, 99999)}"

        students.append({
            "id": sid,
            "name": full_name,
            "gender": gender,
            "course": course,
            "year": year,
            "cgpa": cgpa,
            "credits": credits_val,
            "attendance": attd,
            "lms_score": lms,
            "father": father_name,
            "mother": mother_name,
            "mother_tongue": mother_tongue,
            "place": place,
            "region": region,
            "country": "India",
            "email": slug,
            "phone": phone
        })

    # Compute exact explainable risk score for every student
    for s in students:
        cgpa_scaled = s["cgpa"] * 10.0
        engagement = (s["attendance"] * 0.40) + (cgpa_scaled * 0.35) + (s["lms_score"] * 0.25)
        s["risk"] = round(max(0, min(100, 100.0 - engagement)))

    return students


# -------------------------------------------------------------
# 3. GENERATE 50 FACULTY & MENTORS
# -------------------------------------------------------------

FACULTY_DESIGNATIONS = [
    "Professor & Head", "Associate Professor", "Assistant Professor (Sr. Grade)",
    "Assistant Professor", "Dean of Computing", "Professor of Practice"
]

SPECIALIZATIONS_FACULTY = [
    "Database Systems & Query Optimization", "Operating Systems & Distributed Computing",
    "Computer Networks & Cyber Infrastructure", "Agile Software Architecture & DevOps",
    "Discrete Mathematics & Quantum Algorithms", "Deep Learning & Natural Language Processing",
    "Cloud Architecture & High-Performance Computing", "Compiler Design & Code Optimization",
    "Full-Stack Web Development & Microservices", "Data Warehousing & Business Intelligence"
]

SPECIALIZATIONS_MENTORS = [
    "Academic Performance & Retention Counseling", "Attendance Intervention & Habit Coaching",
    "Remedial Study Planning & Subject Tutoring", "Career Pathways & Placement Readiness",
    "Emotional Wellness & Academic Resilience", "Peer Mentoring Group Leadership",
    "First-Year Foundation & Bridge Coaching", "Advanced Project & Research Mentorship"
]

EXTRA_ROLES = [
    "Class Teacher, 2nd Year Coordinator", "Exam Cell Incharge", "NBA / NAAC Criteria Lead",
    "Department Placement Coordinator", "Academic Audit Committee Member", "Hostel Warden & Counselor",
    "Lab Incharge (Data Systems)", "Student Grievance Redressal Lead", "Curriculum Revision Incharge"
]

def generate_faculty_and_mentors():
    faculty = []
    mentors = []

    # 1. Faculty 001 - 030 (30 total)
    # Guaranteed initial 3 faculty
    faculty.append({
        "id": "FAC001",
        "display_name": "Dr. Ramesh Kumar",
        "role": "faculty",
        "subjects": "CS201,CS202",
        "department": "CSE",
        "assigned_year": "2nd Year",
        "specialization": "Database Systems & Mentorship",
        "extra_roles": "Class Teacher, 2nd Year Coordinator",
        "email": "dr.ramesh@vignan.ac.in",
        "phone": "+91 90000 11111"
    })
    faculty.append({
        "id": "FAC002",
        "display_name": "Dr. Priya Sharma",
        "role": "faculty",
        "subjects": "CS203,CS204",
        "department": "CSE",
        "assigned_year": "3rd Year",
        "specialization": "Software Engineering & Agile Methodologies",
        "extra_roles": "Exam Cell Incharge",
        "email": "dr.priya@vignan.ac.in",
        "phone": "+91 90000 22222"
    })
    faculty.append({
        "id": "FAC003",
        "display_name": "Prof. Venkat Rao",
        "role": "faculty",
        "subjects": "MA201",
        "department": "Mathematics",
        "assigned_year": "1st Year",
        "specialization": "Discrete Math & Probability Models",
        "extra_roles": "HOD Mathematics",
        "email": "prof.venkat@vignan.ac.in",
        "phone": "+91 90000 33333"
    })

    for i in range(4, 31):
        fid = f"FAC{i:03d}"
        gender = "Male" if random.random() < 0.55 else "Female"
        first = random.choice(FIRST_NAMES_MALE if gender == "Male" else FIRST_NAMES_FEMALE)
        last = random.choice(LAST_NAMES)
        title = "Dr." if random.random() < 0.65 else "Prof."
        name = f"{title} {first} {last}"
        sub_sample = random.sample(["CS201", "CS202", "CS203", "CS204", "MA201", "CS301", "CS302", "CS303"], k=random.randint(1, 2))
        subjects_str = ",".join(sub_sample)
        dept = random.choice(["CSE", "CSE (AI & ML)", "Data Science", "Mathematics", "IT"])
        year = random.choice(["1st Year", "2nd Year", "3rd Year", "4th Year"])
        spec = random.choice(SPECIALIZATIONS_FACULTY)
        extra = random.choice(EXTRA_ROLES) if random.random() < 0.7 else None
        email = f"{first.lower()}.{last.lower()[:4]}.fac@vignan.ac.in"
        phone = f"+91 90000 {i:05d}"

        faculty.append({
            "id": fid,
            "display_name": name,
            "role": "faculty",
            "subjects": subjects_str,
            "department": dept,
            "assigned_year": year,
            "specialization": spec,
            "extra_roles": extra,
            "email": email,
            "phone": phone
        })

    # 2. Mentors 001 - 020 (20 total)
    mentors.append({
        "id": "MEN001",
        "display_name": "Prof. Sunitha Devi",
        "role": "mentor",
        "subjects": "CS201,CS203",
        "department": "CSE",
        "assigned_year": "2nd Year",
        "specialization": "Academic Mentorship & Core Programming",
        "extra_roles": "Senior Counseling Lead",
        "email": "prof.sunitha@vignan.ac.in",
        "phone": "+91 90000 44444"
    })
    mentors.append({
        "id": "MEN002",
        "display_name": "Dr. Anil Kumar",
        "role": "mentor",
        "subjects": "CS202,CS204",
        "department": "CSE",
        "assigned_year": "1st Year",
        "specialization": "Attendance Recovery & Foundation Mentoring",
        "extra_roles": "Hostel Discipline Counselor",
        "email": "dr.anil@vignan.ac.in",
        "phone": "+91 90000 55555"
    })

    for i in range(3, 21):
        mid = f"MEN{i:03d}"
        gender = "Male" if random.random() < 0.5 else "Female"
        first = random.choice(FIRST_NAMES_MALE if gender == "Male" else FIRST_NAMES_FEMALE)
        last = random.choice(LAST_NAMES)
        title = "Dr." if random.random() < 0.55 else "Prof."
        name = f"{title} {first} {last}"
        sub_sample = random.sample(["CS201", "CS202", "CS203", "CS204", "MA201"], k=random.randint(1, 2))
        subjects_str = ",".join(sub_sample)
        dept = random.choice(["CSE", "Student Counseling Cell", "Academic Welfare", "CSE (AI & ML)"])
        year = random.choice(["1st Year", "2nd Year", "3rd Year"])
        spec = random.choice(SPECIALIZATIONS_MENTORS)
        extra = random.choice(["Senior Mentor", "Year Incharge", "Peer Tutor Mentor", None])
        email = f"{first.lower()}.{last.lower()[:4]}.men@vignan.ac.in"
        phone = f"+91 90000 {50000 + i}"

        mentors.append({
            "id": mid,
            "display_name": name,
            "role": "mentor",
            "subjects": subjects_str,
            "department": dept,
            "assigned_year": year,
            "specialization": spec,
            "extra_roles": extra,
            "email": email,
            "phone": phone
        })

    return faculty, mentors


# -------------------------------------------------------------
# 4. GENERATE 10 ADMINS
# -------------------------------------------------------------

def generate_admins():
    admins = [
        {
            "id": "admin",
            "password": "admin123",
            "role": "admin",
            "display_name": "System Administrator",
            "email": "admin@vignan.ac.in",
            "phone": "+91 90000 00001",
            "department": "IT Infrastructure & Administration",
            "specialization": "Full Platform Control & Institutional AI Operations",
            "extra_roles": "Chief System Administrator"
        },
        {
            "id": "ADM001",
            "password": "ADM001",
            "role": "admin",
            "display_name": "Dr. P. Venkateswarlu",
            "email": "dean.academics@vignan.ac.in",
            "phone": "+91 90000 00002",
            "department": "Academic Affairs",
            "specialization": "Institutional Academic Policy & Curriculum Standards",
            "extra_roles": "Dean of Academic Affairs"
        },
        {
            "id": "ADM002",
            "password": "ADM002",
            "role": "admin",
            "display_name": "Dr. K. Radhika",
            "email": "hod.cse@vignan.ac.in",
            "phone": "+91 90000 00003",
            "department": "CSE",
            "specialization": "Departmental Governance & Faculty Allocations",
            "extra_roles": "Head of Department (CSE)"
        },
        {
            "id": "ADM003",
            "password": "ADM003",
            "role": "admin",
            "display_name": "Dr. M. Sridhar",
            "email": "controller.exams@vignan.ac.in",
            "phone": "+91 90000 00004",
            "department": "Examination Cell",
            "specialization": "Grading Audits, External Exams & Moderation",
            "extra_roles": "Controller of Examinations"
        },
        {
            "id": "ADM004",
            "password": "ADM004",
            "role": "admin",
            "display_name": "Prof. S. Nageswara Rao",
            "email": "director.welfare@vignan.ac.in",
            "phone": "+91 90000 00005",
            "department": "Student Welfare",
            "specialization": "Campus Discipline, Student Wellness & Scholarships",
            "extra_roles": "Director of Student Welfare"
        },
        {
            "id": "ADM005",
            "password": "ADM005",
            "role": "admin",
            "display_name": "Dr. B. Chandrasekhar",
            "email": "iqac.lead@vignan.ac.in",
            "phone": "+91 90000 00006",
            "department": "Internal Quality Assurance Cell (IQAC)",
            "specialization": "Institutional Quality Metrics & NAAC/NBA Compliance",
            "extra_roles": "Chief IQAC Coordinator"
        },
        {
            "id": "ADM006",
            "password": "ADM006",
            "role": "admin",
            "display_name": "Prof. V. Padmavathi",
            "email": "head.mentoring@vignan.ac.in",
            "phone": "+91 90000 00007",
            "department": "Counseling & Mentorship Directorate",
            "specialization": "Multi-Tier Academic Counseling Supervision",
            "extra_roles": "Chief Mentorship Director"
        },
        {
            "id": "ADM007",
            "password": "ADM007",
            "role": "admin",
            "display_name": "Dr. T. Sudhir",
            "email": "dean.admissions@vignan.ac.in",
            "phone": "+91 90000 00008",
            "department": "Admissions & Records",
            "specialization": "Enrollment Analytics & Student Record Lifecycle",
            "extra_roles": "Dean of Admissions"
        },
        {
            "id": "ADM008",
            "password": "ADM008",
            "role": "admin",
            "display_name": "Prof. G. Ravi Kiran",
            "email": "audit.superintendent@vignan.ac.in",
            "phone": "+91 90000 00009",
            "department": "Academic Audit Cell",
            "specialization": "Continuous Internal Evaluation (CIE) Verification",
            "extra_roles": "Academic Audit Superintendent"
        },
        {
            "id": "ADM009",
            "password": "ADM009",
            "role": "admin",
            "display_name": "Dr. L. Himabindu",
            "email": "accreditation.lead@vignan.ac.in",
            "phone": "+91 90000 00010",
            "department": "Accreditation & Compliance",
            "specialization": "Outcome-Based Education (OBE) Attainment Mapping",
            "extra_roles": "OBE & Compliance Lead"
        }
    ]
    return admins


# -------------------------------------------------------------
# 5. GENERATE SUBJECT MARKS & EXTRACURRICULAR ACTIVITIES
# -------------------------------------------------------------

def generate_subject_marks(students):
    marks = []
    # 5 standard 2nd year subjects
    core_subjects = ["CS201", "CS202", "CS203", "CS204", "MA201"]

    for s in students:
        base_attd = s["attendance"]
        base_cgpa = s["cgpa"]

        for code in core_subjects:
            # Add realistic variance per subject
            s_attd = max(35, min(100, int(base_attd + random.randint(-8, 8))))
            
            # Map CGPA (0-10) to internal (0-30) and external (0-70)
            perf_factor = (base_cgpa / 10.0) + (random.uniform(-0.10, 0.10))
            perf_factor = max(0.35, min(0.98, perf_factor))

            internal = int(round(perf_factor * 30))
            external = int(round(perf_factor * 70))
            assignment = int(round(perf_factor * 100))

            total = internal + external
            if total >= 90:
                grade = "A+"
            elif total >= 80:
                grade = "A"
            elif total >= 70:
                grade = "B+"
            elif total >= 60:
                grade = "B"
            elif total >= 50:
                grade = "C"
            elif total >= 40:
                grade = "D"
            else:
                grade = "F"

            marks.append({
                "student_id": s["id"],
                "subject_code": code,
                "attendance": s_attd,
                "internal_marks": internal,
                "external_marks": external,
                "assignment_score": assignment,
                "grade": grade
            })

    return marks


def generate_student_activities(students):
    activities = []
    for s in students:
        # Each student participates in 1 to 3 clubs/activities
        k = random.randint(1, 3)
        chosen = random.sample(EXTRACURRICULARS_POOL, k=k)
        for act in chosen:
            act_id = act[0]
            role = "Participant"
            if s["cgpa"] > 8.5 and random.random() < 0.35:
                role = "Lead / Coordinator" if random.random() < 0.5 else "Winner / 1st Place"
            elif random.random() < 0.2:
                role = "Active Member"

            activities.append({
                "student_id": s["id"],
                "activity_id": act_id,
                "role": role,
                "date_joined": "2026-01-15",
                "notes": f"Active contribution in {act[1]}"
            })
    return activities


# -------------------------------------------------------------
# 6. GENERATE INTERVENTIONS, NOTIFICATIONS, SIGNUPS
# -------------------------------------------------------------

def generate_interventions(students, faculty, mentors):
    interventions = []
    
    # Priority students (High & Med Risk)
    at_risk_students = [s for s in students if s["risk"] >= 40]
    
    actions = [
        "1-on-1 Academic Counseling Session",
        "Attendance Recovery & Root-Cause Review",
        "Remedial Quiz & Assignment Support",
        "LMS Inactivity & Engagement Diagnostic",
        "Peer-Assisted Study Group Assignment",
        "Parent-Teacher Counseling Conference",
        "Mid-Term Performance Remediation Plan"
    ]

    mentor_ids = [m["id"] for m in mentors]
    fac_ids = [f["id"] for f in faculty]

    today = datetime.date.today()

    # Create 35 representative interventions
    for i in range(1, 36):
        s = at_risk_students[i % len(at_risk_students)]
        action = random.choice(actions)
        subject_code = random.choice(["CS201", "CS202", "CS203", "CS204", "MA201"])
        mentor_id = random.choice(mentor_ids[:8]) # Focus on top mentors
        created_by = random.choice([mentor_id, random.choice(fac_ids[:6]), "admin"])
        
        days_ago = random.randint(2, 45)
        int_date = (today - datetime.timedelta(days=days_ago)).isoformat()
        
        # Balance statuses across all review pipelines
        if i <= 15:
            # Completed
            status = "Completed"
            req_by = mentor_id
            req_notes = "Conducted rigorous session. Student submitted backlogged assignments and committed to 85% attendance."
            req_at = (today - datetime.timedelta(days=days_ago-2)).isoformat()
            rev_by = "admin" if random.random() < 0.5 else "FAC001"
            rev_at = (today - datetime.timedelta(days=days_ago-3)).isoformat()
            rej_reason = None
        elif i <= 23:
            # Completion Requested (Pending Enquiries)
            status = "Completion Requested"
            req_by = mentor_id
            req_notes = "Session completed successfully. Remedial test taken with 24/30 score. Requesting sign-off."
            req_at = (today - datetime.timedelta(days=random.randint(1, 4))).isoformat()
            rev_by = None
            rev_at = None
            rej_reason = None
        elif i <= 28:
            # Revision Needed
            status = "Revision Needed"
            req_by = mentor_id
            req_notes = "Initial meeting conducted. Student promised improvement."
            req_at = (today - datetime.timedelta(days=random.randint(5, 10))).isoformat()
            rev_by = "FAC001"
            rev_at = (today - datetime.timedelta(days=random.randint(1, 4))).isoformat()
            rej_reason = "Attendance did not show measurable uptick in week 3. Please conduct a follow-up test and verify LMS log before re-submitting."
        elif i <= 32:
            status = "In Progress"
            req_by = None
            req_notes = None
            req_at = None
            rev_by = None
            rev_at = None
            rej_reason = None
        else:
            status = "Pending"
            req_by = None
            req_notes = None
            req_at = None
            rev_by = None
            rev_at = None
            rej_reason = None

        interventions.append({
            "id": i,
            "student_id": s["id"],
            "date": int_date,
            "action": action,
            "status": status,
            "notes": f"Intervention targeted for {s['name']} regarding {subject_code} performance slump.",
            "urgency": "Critical" if s["risk"] > 60 else ("High" if s["risk"] > 45 else "Moderate"),
            "subject_code": subject_code,
            "mentor_id": mentor_id,
            "created_by": created_by,
            "session_time": "10:30 AM" if i % 2 == 0 else "02:30 PM",
            "location": "Mentorship Cabin 204, CSE Block" if i % 2 == 0 else "Google Meet (Online)",
            "completion_requested_by": req_by,
            "completion_request_notes": req_notes,
            "completion_requested_at": req_at,
            "reviewed_by": rev_by,
            "reviewed_at": rev_at,
            "rejection_reason": rej_reason
        })

    return interventions


# -------------------------------------------------------------
# 7. SEED SQLITE DATABASE
# -------------------------------------------------------------

def seed_database(students, faculty, mentors, admins, subject_marks, student_activities, interventions):
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Execute full DDL
    c.executescript("""
        CREATE TABLE students (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            gender TEXT,
            course TEXT,
            year TEXT,
            cgpa REAL,
            credits INTEGER,
            attendance INTEGER,
            lms_score INTEGER,
            risk INTEGER,
            father TEXT,
            mother TEXT,
            mother_tongue TEXT,
            place TEXT,
            region TEXT,
            country TEXT,
            email TEXT,
            phone TEXT
        );

        CREATE TABLE users (
            id TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            role TEXT NOT NULL,
            display_name TEXT,
            linked_student_id TEXT,
            subjects TEXT,
            extra_roles TEXT,
            email TEXT,
            phone TEXT,
            status TEXT DEFAULT 'Active',
            department TEXT DEFAULT 'CSE',
            assigned_year TEXT DEFAULT '2nd Year',
            specialization TEXT DEFAULT 'Academic Counseling'
        );

        CREATE TABLE signup_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL,
            display_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT,
            subjects TEXT,
            extra_roles TEXT,
            department TEXT DEFAULT 'CSE',
            assigned_year TEXT DEFAULT '2nd Year',
            specialization TEXT DEFAULT 'Academic Counseling',
            status TEXT DEFAULT 'Pending',
            rejection_reason TEXT,
            created_at TEXT NOT NULL,
            reviewed_at TEXT
        );

        CREATE TABLE email_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            recipient TEXT NOT NULL,
            subject TEXT NOT NULL,
            body_html TEXT NOT NULL,
            body_text TEXT,
            email_type TEXT NOT NULL,
            sent_at TEXT NOT NULL
        );

        CREATE TABLE subjects (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            short_name TEXT,
            credits INTEGER DEFAULT 3,
            year TEXT NOT NULL,
            semester INTEGER DEFAULT 1
        );

        CREATE TABLE subject_marks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            subject_code TEXT NOT NULL,
            attendance INTEGER DEFAULT 0,
            internal_marks INTEGER DEFAULT 0,
            external_marks INTEGER DEFAULT 0,
            assignment_score INTEGER DEFAULT 0,
            grade TEXT DEFAULT 'N/A',
            UNIQUE(student_id, subject_code)
        );

        CREATE TABLE extracurriculars (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT
        );

        CREATE TABLE student_activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            activity_id TEXT NOT NULL,
            role TEXT DEFAULT 'Participant',
            date_joined TEXT,
            notes TEXT,
            UNIQUE(student_id, activity_id)
        );

        CREATE TABLE interventions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT,
            date TEXT,
            action TEXT,
            status TEXT DEFAULT 'Pending',
            notes TEXT,
            urgency TEXT DEFAULT 'Moderate',
            subject_code TEXT,
            mentor_id TEXT,
            completed_date TEXT,
            created_by TEXT,
            session_time TEXT DEFAULT '10:00 AM',
            location TEXT DEFAULT 'Mentorship Cabin 204',
            completion_requested_by TEXT,
            completion_request_notes TEXT,
            completion_requested_at TEXT,
            rejection_reason TEXT,
            reviewed_by TEXT,
            reviewed_at TEXT
        );

        CREATE TABLE notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            message TEXT NOT NULL,
            type TEXT DEFAULT 'info',
            date TEXT NOT NULL,
            student_id TEXT,
            read INTEGER DEFAULT 0
        );

        CREATE TABLE agent_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT,
            student_name TEXT,
            timestamp TEXT,
            diagnosis TEXT,
            actions_taken TEXT
        );

        CREATE TABLE settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    """)

    # 1. Insert Students
    for s in students:
        c.execute("""
            INSERT INTO students (id, name, gender, course, year, cgpa, credits, attendance, lms_score, risk, father, mother, mother_tongue, place, region, country, email, phone)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            s["id"], s["name"], s["gender"], s["course"], s["year"], s["cgpa"], s["credits"],
            s["attendance"], s["lms_score"], s["risk"], s["father"], s["mother"], s["mother_tongue"],
            s["place"], s["region"], s["country"], s["email"], s["phone"]
        ))

    # 2. Insert Users (Admins, Faculty, Mentors, Students)
    for a in admins:
        c.execute("""
            INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?, ?, ?, ?, NULL, NULL, ?, ?, ?, 'Active', ?, 'All Years', ?)
        """, (a["id"], a["password"], a["role"], a["display_name"], a["extra_roles"], a["email"], a["phone"], a["department"], a["specialization"]))

    for f in faculty:
        c.execute("""
            INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?, ?, ?, ?, NULL, ?, ?, ?, ?, 'Active', ?, ?, ?)
        """, (f["id"], f["id"], f["role"], f["display_name"], f["subjects"], f["extra_roles"], f["email"], f["phone"], f["department"], f["assigned_year"], f["specialization"]))

    for m in mentors:
        c.execute("""
            INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?, ?, ?, ?, NULL, ?, ?, ?, ?, 'Active', ?, ?, ?)
        """, (m["id"], m["id"], m["role"], m["display_name"], m["subjects"], m["extra_roles"], m["email"], m["phone"], m["department"], m["assigned_year"], m["specialization"]))

    for s in students:
        c.execute("""
            INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?, ?, 'student', ?, ?, NULL, NULL, ?, ?, 'Active', ?, ?, 'Student Academic Telemetry')
        """, (s["id"], s["id"], s["name"], s["id"], s["email"], s["phone"], s["course"], s["year"]))

    # 3. Insert Subjects Master
    for sub in SUBJECT_POOL:
        c.execute("""
            INSERT INTO subjects (code, name, short_name, credits, year, semester)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (sub["code"], sub["name"], sub["short_name"], sub["credits"], sub["year"], sub["sem"]))

    # 4. Insert Subject Marks
    for sm in subject_marks:
        c.execute("""
            INSERT INTO subject_marks (student_id, subject_code, attendance, internal_marks, external_marks, assignment_score, grade)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (sm["student_id"], sm["subject_code"], sm["attendance"], sm["internal_marks"], sm["external_marks"], sm["assignment_score"], sm["grade"]))

    # 5. Insert Extracurricular Master
    for act in EXTRACURRICULARS_POOL:
        c.execute("""
            INSERT INTO extracurriculars (id, name, category, description)
            VALUES (?, ?, ?, ?)
        """, (act[0], act[1], act[2], act[3]))

    # 6. Insert Student Activities
    for sa in student_activities:
        c.execute("""
            INSERT INTO student_activities (student_id, activity_id, role, date_joined, notes)
            VALUES (?, ?, ?, ?, ?)
        """, (sa["student_id"], sa["activity_id"], sa["role"], sa["date_joined"], sa["notes"]))

    # 7. Insert Interventions
    for inv in interventions:
        c.execute("""
            INSERT INTO interventions (
                student_id, date, action, status, notes, urgency, subject_code, mentor_id,
                created_by, session_time, location, completion_requested_by, completion_request_notes,
                completion_requested_at, rejection_reason, reviewed_by, reviewed_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            inv["student_id"], inv["date"], inv["action"], inv["status"], inv["notes"],
            inv["urgency"], inv["subject_code"], inv["mentor_id"], inv["created_by"],
            inv["session_time"], inv["location"], inv["completion_requested_by"],
            inv["completion_request_notes"], inv["completion_requested_at"],
            inv["rejection_reason"], inv["reviewed_by"], inv["reviewed_at"]
        ))

    # 8. Insert Signup Requests (Pending, Approved, Declined)
    today = datetime.date.today().isoformat()
    c.execute("""
        INSERT INTO signup_requests (user_id, password, role, display_name, email, phone, subjects, extra_roles, department, assigned_year, specialization, status, created_at)
        VALUES ('FAC031', 'Pass@123', 'faculty', 'Dr. Rajesh Gupta', 'rajesh.gupta@vignan.ac.in', '+91 98765 11223', 'CS201,CS303', 'Lab Incharge', 'CSE', '2nd Year', 'Cloud Microservices & Distributed DBs', 'Pending', ?)
    """, (today,))

    c.execute("""
        INSERT INTO signup_requests (user_id, password, role, display_name, email, phone, subjects, extra_roles, department, assigned_year, specialization, status, created_at)
        VALUES ('MEN021', 'Pass@123', 'mentor', 'Prof. Ananya Sen', 'ananya.sen@vignan.ac.in', '+91 98765 44332', 'CS101,CS202', 'Counselor', 'Student Welfare', '1st Year', 'First Year Transition Counseling', 'Pending', ?)
    """, (today,))

    c.execute("""
        INSERT INTO signup_requests (user_id, password, role, display_name, email, phone, subjects, extra_roles, department, assigned_year, specialization, status, created_at, reviewed_at)
        VALUES ('FAC032', 'Pass@123', 'faculty', 'Dr. Siddharth Roy', 'siddharth.roy@vignan.ac.in', '+91 98765 55667', 'CS301', 'NBA Member', 'CSE', '3rd Year', 'Algorithms & Complexity', 'Approved', ?, ?)
    """, (today, today))

    c.execute("""
        INSERT INTO signup_requests (user_id, password, role, display_name, email, phone, subjects, extra_roles, department, assigned_year, specialization, status, rejection_reason, created_at, reviewed_at)
        VALUES ('FAC033', 'Pass@123', 'faculty', 'Dr. Alok Verma', 'alok.verma@vignan.ac.in', '+91 98765 99887', 'CS204', 'None', 'CSE', '2nd Year', 'General SE', 'Declined', 'Incomplete departmental accreditation documentation and duplicate mobile number provided.', ?, ?)
    """, (today, today))

    # 9. Insert Notifications
    c.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES 
        ('Welcome to EduStudent Sight', 'Institutional Academic Intelligence Platform active. All 100 student telemetry profiles synchronized.', 'info', ?, NULL, 0),
        ('Critical Risk Flag', 'Arjun Patel (25CS005) attendance fell below 65%. 1-on-1 counseling initiated.', 'danger', ?, '25CS005', 0),
        ('Mentorship Review Pending', 'Prof. Sunitha Devi requested completion approval for Y. Hemanth Reddy (25CS002).', 'warning', ?, '25CS002', 0),
        ('Intervention Completed', 'Dr. Ramesh Kumar verified attendance improvement for V. Sri Udbhav (25CS001).', 'success', ?, '25CS001', 1)
    """, (today, today, today, today))

    # 10. Default System Settings
    settings_dict = {
        "risk_threshold": "60",
        "safe_risk_threshold": "50",
        "high_risk_threshold": "40",
        "critical_risk_threshold": "65",
        "attendance_threshold": "75",
        "risk_cgpa_threshold": "7.5",
        "lms_threshold": "60",
        "assignment_threshold": "70",
        "ai_provider": "gemini",
        "model_name": "gemini-3.5-flash",
        "api_base_url": ""
    }
    for k, v in settings_dict.items():
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, v))

    conn.commit()
    conn.close()
    print(f"✓ SQLite Database created and seeded successfully at: {DB_PATH}")


# -------------------------------------------------------------
# 8. EXPORT CLEAN CSV AND EXCEL SPREADSHEETS
# -------------------------------------------------------------

def export_csv_and_excel(students, faculty, mentors, admins, subject_marks, student_activities, interventions):
    
    # 1. Students CSV
    students_csv_path = os.path.join(DATA_DIR, "students_100.csv")
    with open(students_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "name", "gender", "course", "year", "cgpa", "credits", "attendance",
            "lms_score", "risk", "father", "mother", "mother_tongue", "place", "region", "country", "email", "phone"
        ])
        writer.writeheader()
        writer.writerows(students)
    print(f"✓ Exported: {students_csv_path}")

    # 2. Faculty & Mentors CSV
    fac_men_csv_path = os.path.join(DATA_DIR, "faculty_and_mentors_50.csv")
    with open(fac_men_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "display_name", "role", "department", "assigned_year", "subjects", "specialization", "extra_roles", "email", "phone"
        ])
        writer.writeheader()
        writer.writerows(faculty + mentors)
    print(f"✓ Exported: {fac_men_csv_path}")

    # 3. Admins CSV
    admins_csv_path = os.path.join(DATA_DIR, "admins_10.csv")
    with open(admins_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "display_name", "role", "department", "specialization", "extra_roles", "email", "phone"
        ])
        writer.writeheader()
        for a in admins:
            row = {k: a[k] for k in writer.fieldnames}
            writer.writerow(row)
    print(f"✓ Exported: {admins_csv_path}")

    # 4. Master Credentials CSV
    creds_csv_path = os.path.join(DATA_DIR, "master_credentials.csv")
    with open(creds_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["User ID", "Password", "Role", "Display Name", "Email", "Department / Program", "Specialization / Responsibility"])
        for a in admins:
            writer.writerow([a["id"], a["password"], a["role"].upper(), a["display_name"], a["email"], a["department"], a["extra_roles"]])
        for f_item in faculty:
            writer.writerow([f_item["id"], f_item["id"], "FACULTY", f_item["display_name"], f_item["email"], f_item["department"], f_item["specialization"]])
        for m in mentors:
            writer.writerow([m["id"], m["id"], "MENTOR", m["display_name"], m["email"], m["department"], m["specialization"]])
        for s in students:
            writer.writerow([s["id"], s["id"], "STUDENT", s["name"], s["email"], f"{s['course']} {s['year']}", f"Risk: {s['risk']}% | CGPA: {s['cgpa']}"])
    print(f"✓ Exported: {creds_csv_path}")

    # 5. Subject Marks CSV
    marks_csv_path = os.path.join(DATA_DIR, "subject_marks_500.csv")
    with open(marks_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["student_id", "subject_code", "attendance", "internal_marks", "external_marks", "assignment_score", "grade"])
        writer.writeheader()
        writer.writerows(subject_marks)
    print(f"✓ Exported: {marks_csv_path}")

    # 6. Interventions CSV
    inv_csv_path = os.path.join(DATA_DIR, "interventions.csv")
    with open(inv_csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "student_id", "date", "action", "status", "urgency", "subject_code",
            "mentor_id", "created_by", "session_time", "location", "completion_requested_by",
            "completion_request_notes", "reviewed_by", "rejection_reason"
        ])
        writer.writeheader()
        for item in interventions:
            row = {k: item.get(k) for k in writer.fieldnames}
            writer.writerow(row)
    print(f"[OK] Exported: {inv_csv_path}")

    # 7. Multi-Sheet Professional Excel Workbook (.xlsx)
    if HAS_OPENPYXL:
        try:
            excel_path = os.path.join(DATA_DIR, "EduStudent_Sight_Master_Database.xlsx")
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=10)
            border_thin = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            def style_sheet(ws, title, headers, rows):
                ws.title = title
                ws.append(headers)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for row in rows:
                    ws.append(row)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = data_font
                        cell.border = border_thin
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="left")

                # Auto-adjust column widths
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # Sheet 1: Students (100)
            ws_stu = wb.create_sheet()
            stu_headers = ["Student ID", "Full Name", "Gender", "Course", "Year", "CGPA", "Credits", "Attendance %", "LMS Score %", "Risk %", "Mother Tongue", "City", "Region", "Email", "Mobile"]
            stu_rows = [
                [s["id"], s["name"], s["gender"], s["course"], s["year"], s["cgpa"], s["credits"], s["attendance"], s["lms_score"], s["risk"], s["mother_tongue"], s["place"], s["region"], s["email"], s["phone"]]
                for s in students
            ]
            style_sheet(ws_stu, "Students (100)", stu_headers, stu_rows)

            # Sheet 2: Faculty & Mentors (50)
            ws_fac = wb.create_sheet()
            fac_headers = ["Faculty/Mentor ID", "Display Name", "Role", "Department", "Assigned Year", "Subject Codes", "Specialization Area", "Institutional Responsibility", "Email", "Phone"]
            fac_rows = [
                [u["id"], u["display_name"], u["role"].capitalize(), u["department"], u["assigned_year"], u["subjects"], u["specialization"], u["extra_roles"] or "None", u["email"], u["phone"]]
                for u in (faculty + mentors)
            ]
            style_sheet(ws_fac, "Faculty & Mentors (50)", fac_headers, fac_rows)

            # Sheet 3: Administrators (10)
            ws_adm = wb.create_sheet()
            adm_headers = ["Admin ID", "Display Name", "Designation / Role", "Department / Directorate", "Specialization Scope", "Institutional Email", "Mobile Number"]
            adm_rows = [
                [a["id"], a["display_name"], a["extra_roles"], a["department"], a["specialization"], a["email"], a["phone"]]
                for a in admins
            ]
            style_sheet(ws_adm, "System Admins (10)", adm_headers, adm_rows)

            # Sheet 4: Subject Marks (500)
            ws_mrk = wb.create_sheet()
            mrk_headers = ["Student ID", "Subject Code", "Attendance %", "Internal Marks (30)", "External Marks (70)", "Assignment (100)", "Letter Grade"]
            mrk_rows = [
                [m["student_id"], m["subject_code"], m["attendance"], m["internal_marks"], m["external_marks"], m["assignment_score"], m["grade"]]
                for m in subject_marks
            ]
            style_sheet(ws_mrk, "Subject Marks (500)", mrk_headers, mrk_rows)

            # Sheet 5: Interventions (35)
            ws_inv = wb.create_sheet()
            inv_headers = ["ID", "Student ID", "Session Date", "Action Strategy", "Status", "Urgency", "Subject", "Mentor ID", "Created By", "Time", "Location", "Reviewer Status Notes"]
            inv_rows = [
                [i["id"], i["student_id"], i["date"], i["action"], i["status"], i["urgency"], i["subject_code"], i["mentor_id"], i["created_by"], i["session_time"], i["location"], i["rejection_reason"] or i["completion_request_notes"] or "Nominal"]
                for i in interventions
            ]
            style_sheet(ws_inv, "Interventions Radar (35)", inv_headers, inv_rows)

            # Sheet 6: Master Credentials Access Sheet
            ws_crd = wb.create_sheet()
            crd_headers = ["User ID", "Login Password", "Access Level", "Display Name", "Official Email", "Assigned Unit / Cohort", "Telemetry Overview"]
            crd_rows = []
            for a in admins:
                crd_rows.append([a["id"], a["password"], "ADMINISTRATOR", a["display_name"], a["email"], a["department"], a["extra_roles"]])
            for f_item in faculty:
                crd_rows.append([f_item["id"], f_item["id"], "FACULTY", f_item["display_name"], f_item["email"], f_item["department"], f_item["specialization"]])
            for m in mentors:
                crd_rows.append([m["id"], m["id"], "MENTOR", m["display_name"], m["email"], m["department"], m["specialization"]])
            for s in students:
                crd_rows.append([s["id"], s["id"], "STUDENT", s["name"], s["email"], f"{s['course']} {s['year']}", f"Risk: {s['risk']}% | CGPA: {s['cgpa']} | Attd: {s['attendance']}%"])
            style_sheet(ws_crd, "Master Credentials (160)", crd_headers, crd_rows)

            wb.save(excel_path)
            print(f"[OK] Exported Multi-Tab Excel Workbook: {excel_path}")
        except Exception as e:
            print(f"[Warning] Excel export skipped: {e}")


# -------------------------------------------------------------
# MAIN EXECUTION
# -------------------------------------------------------------

def main():
    print("=" * 60)
    print("Generating Complete Demo Data for EduStudent Sight...")
    print("=" * 60)

    students = generate_students(100)
    faculty, mentors = generate_faculty_and_mentors()
    admins = generate_admins()
    subject_marks = generate_subject_marks(students)
    student_activities = generate_student_activities(students)
    interventions = generate_interventions(students, faculty, mentors)

    print(f"[OK] Generated {len(students)} Student profiles")
    print(f"[OK] Generated {len(faculty)} Faculty members")
    print(f"[OK] Generated {len(mentors)} Mentors")
    print(f"[OK] Generated {len(admins)} Administrators")
    print(f"[OK] Generated {len(subject_marks)} Subject Marks records")
    print(f"[OK] Generated {len(student_activities)} Student Activity memberships")
    print(f"[OK] Generated {len(interventions)} Interventions & Enquiries")

    print("\nSeeding SQLite Database...")
    seed_database(students, faculty, mentors, admins, subject_marks, student_activities, interventions)

    print("\nExporting Datasets to 'data/'...")
    export_csv_and_excel(students, faculty, mentors, admins, subject_marks, student_activities, interventions)

    print("\n" + "=" * 60)
    print("DEMO DATASET CREATION COMPLETE!")
    print("=" * 60)

if __name__ == "__main__":
    main()
