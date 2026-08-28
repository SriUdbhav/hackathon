# =====================================================
# APP.PY
# Flask REST API — Auth, CRUD, UAC, Subjects, Activities,
# Intervention Lifecycle, Autonomous Agent, Settings
# =====================================================

from flask import Flask, request, jsonify
from flask_cors import CORS
import db
import ai_engine
from agent import AcademicInterventionAgent
import datetime
import csv
import io
import os


# =====================================================
# .ENV LOADER (no python-dotenv dependency required)
# =====================================================
def _load_dotenv(path=None):
    """Load key=value pairs from a .env file into os.environ."""
    candidates = [
        path,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
        "/etc/secrets/.env",
        os.path.join(os.getcwd(), ".env"),
        os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
    ]
    for p in candidates:
        if p and os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line or line.startswith("#") or "=" not in line:
                            continue
                        key, _, value = line.partition("=")
                        key = key.strip()
                        value = value.strip().strip('"').strip("'")
                        if key and value and key not in os.environ:
                            os.environ[key] = value
            except Exception as e:
                print(f"[Warning] Failed to load .env from {p}: {e}")

_load_dotenv()

app = Flask(__name__)
CORS(app)

db.init_db()
agent = AcademicInterventionAgent()


@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "status": "Online",
        "service": "EduStudent Sight - Autonomous Agentic AI Platform",
        "timestamp": datetime.datetime.now().isoformat()
    })


# =====================================================
# 1. AUTHENTICATION & USER MANAGEMENT
# =====================================================

def _log_email(recipient, subject, body_html, body_text, email_type):
    """Logs sent email notification to database audit trail."""
    try:
        conn = db.get_db_connection()
        conn.execute("""
            INSERT INTO email_logs (recipient, subject, body_html, body_text, email_type, sent_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (recipient, subject, body_html, body_text, email_type, datetime.datetime.now().isoformat()))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Email Log Error] {e}")


@app.route("/api/login", methods=["POST"])
def login():
    """Case-insensitive login using raw user ID or email."""
    data = request.get_json() or {}
    user_id = data.get("email", data.get("id", "")).strip()
    password = data.get("password", "")

    conn = db.get_db_connection()
    # Check if there is a pending signup request for this user ID
    pending = conn.execute("SELECT * FROM signup_requests WHERE (LOWER(user_id) = LOWER(?) OR LOWER(email) = LOWER(?)) AND status = 'Pending'", (user_id, user_id)).fetchone()
    if pending:
        conn.close()
        return jsonify({
            "success": False,
            "message": "Your registration request is currently pending Administrator review. You will receive an email once approved."
        }), 403

    rejected = conn.execute("SELECT * FROM signup_requests WHERE (LOWER(user_id) = LOWER(?) OR LOWER(email) = LOWER(?)) AND status IN ('Rejected', 'Declined')", (user_id, user_id)).fetchone()
    if rejected:
        conn.close()
        return jsonify({
            "success": False,
            "message": f"Your registration request was declined by the administrator. Reason: {rejected['rejection_reason'] or 'Contact department administration.'}"
        }), 403

    # Lookup active user by ID or Email
    user = conn.execute("SELECT * FROM users WHERE LOWER(id) = LOWER(?) OR LOWER(COALESCE(email,'')) = LOWER(?)", (user_id, user_id)).fetchone()
    conn.close()

    if user and user["password"] == password:
        return jsonify({
            "success": True,
            "id": user["id"],
            "role": user["role"],
            "display_name": user["display_name"],
            "linked_student_id": user["linked_student_id"],
            "subjects": user["subjects"],
            "extra_roles": user["extra_roles"],
            "email": user["email"] if "email" in user.keys() else None,
            "phone": user["phone"] if "phone" in user.keys() else None
        })

    return jsonify({"success": False, "message": "Invalid ID or password."}), 401


@app.route("/api/users/password", methods=["PUT"])
def change_password():
    """Change password — requires current password verification."""
    data = request.get_json() or {}
    user_id = data.get("id", "").strip()
    current_pw = data.get("current_password", "")
    new_pw = data.get("new_password", "")

    if not user_id or not new_pw:
        return jsonify({"success": False, "message": "User ID and new password required."}), 400

    conn = db.get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE LOWER(id) = LOWER(?)", (user_id,)).fetchone()

    if not user or user["password"] != current_pw:
        conn.close()
        return jsonify({"success": False, "message": "Current password is incorrect."}), 403

    conn.execute("UPDATE users SET password = ? WHERE LOWER(id) = LOWER(?)", (new_pw, user_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Password updated successfully."})


@app.route("/api/users", methods=["GET"])
def get_all_users():
    """Admin only: List all user accounts."""
    conn = db.get_db_connection()
    rows = conn.execute("SELECT id, role, display_name, linked_student_id, subjects, extra_roles, email, phone FROM users").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/users", methods=["POST"])
def create_user():
    """Admin only: Create a new faculty or mentor user directly with department and assigned cohort."""
    data = request.get_json() or {}
    user_id = data.get("id", "").strip()
    if not user_id:
        return jsonify({"success": False, "message": "User ID required."}), 400

    conn = db.get_db_connection()
    try:
        conn.execute("""
            INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            user_id,
            data.get("password", user_id),
            data.get("role", "mentor"),
            data.get("display_name", user_id),
            data.get("linked_student_id"),
            data.get("subjects", ""),
            data.get("extra_roles", ""),
            data.get("email", ""),
            data.get("phone", ""),
            data.get("status", "Active"),
            data.get("department", "CSE"),
            data.get("assigned_year", "2nd Year"),
            data.get("specialization", "Academic Counseling")
        ))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": f"User {user_id} created successfully."})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/users/<user_id>", methods=["DELETE"])
def delete_user(user_id):
    """Admin only: Delete a user account."""
    conn = db.get_db_connection()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"User {user_id} deleted."})


# =====================================================
# 1C. USER PROFILE ENDPOINT
# =====================================================

@app.route("/api/profile/<user_id>", methods=["GET"])
def get_profile(user_id):
    """Get complete profile for a user, merging users + students tables."""
    conn = db.get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE LOWER(id) = LOWER(?) OR LOWER(display_name) = LOWER(?) OR LOWER(COALESCE(email,'')) = LOWER(?)", (user_id, user_id, user_id)).fetchone()
    if not user:
        conn.close()
        return jsonify({"success": False, "message": "User not found."}), 404

    profile = {
        "id": user["id"],
        "display_name": user["display_name"],
        "role": user["role"],
        "email": user["email"] if "email" in user.keys() else None,
        "phone": user["phone"] if "phone" in user.keys() else None,
        "subjects": user["subjects"],
        "extra_roles": user["extra_roles"],
        "linked_student_id": user["linked_student_id"],
    }

    # If the user is a student with a linked student record, merge academic data
    linked_id = user["linked_student_id"] or user["id"]
    student = conn.execute("SELECT * FROM students WHERE LOWER(id) = LOWER(?)", (linked_id,)).fetchone()
    if student:
        profile["student_data"] = {
            "id": student["id"],
            "name": student["name"],
            "gender": student["gender"],
            "course": student["course"],
            "year": student["year"],
            "cgpa": student["cgpa"],
            "credits": student["credits"],
            "attendance": student["attendance"],
            "lms_score": student["lms_score"],
            "risk": student["risk"],
            "father": student["father"],
            "mother": student["mother"],
            "mother_tongue": student["mother_tongue"],
            "place": student["place"],
            "region": student["region"],
            "country": student["country"],
            "email": student["email"] if "email" in student.keys() else None,
            "phone": student["phone"] if "phone" in student.keys() else None,
        }

    conn.close()
    return jsonify({"success": True, "profile": profile})


# =====================================================
# 1B. SIGNUP REQUESTS & ADMIN APPROVAL WORKFLOW
# =====================================================

@app.route("/api/signup-requests", methods=["POST"])
def submit_signup_request():
    """Submit a registration application for Faculty or Mentor account."""
    data = request.get_json() or {}
    user_id = data.get("id", data.get("user_id", "")).strip()
    name = data.get("display_name", data.get("name", "")).strip()
    role = data.get("role", "faculty").strip().lower()
    email = data.get("email", "").strip()
    phone = data.get("phone", "").strip()
    password = data.get("password", "").strip()
    subjects = data.get("subjects", "").strip()
    extra_roles = data.get("extra_roles", "").strip()

    if not user_id or not name or not email or not password:
        return jsonify({"success": False, "message": "User ID, Full Name, Email, and Password are required."}), 400

    conn = db.get_db_connection()

    # Check if user ID already exists in users
    existing_user = conn.execute("SELECT id FROM users WHERE LOWER(id) = LOWER(?)", (user_id,)).fetchone()
    if existing_user:
        conn.close()
        return jsonify({"success": False, "message": f"User ID '{user_id}' is already registered in the system."}), 400

    # Check if email is already in use
    existing_email = conn.execute("SELECT id FROM users WHERE LOWER(email) = LOWER(?)", (email,)).fetchone()
    if existing_email:
        conn.close()
        return jsonify({"success": False, "message": f"Email address '{email}' is already associated with an existing account."}), 400

    # Check if there is an active pending request
    existing_req = conn.execute("SELECT id FROM signup_requests WHERE (LOWER(user_id) = LOWER(?) OR LOWER(email) = LOWER(?)) AND status = 'Pending'", (user_id, email)).fetchone()
    if existing_req:
        conn.close()
        return jsonify({"success": False, "message": "An application with this ID or Email is already pending Administrator review."}), 400

    try:
        cur = conn.execute("""
            INSERT INTO signup_requests (user_id, password, role, display_name, email, phone, subjects, extra_roles, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?)
        """, (user_id, password, role, name, email, phone, subjects, extra_roles, datetime.date.today().isoformat()))
        req_id = cur.lastrowid

        # Also add a system notification for Admin
        conn.execute("""
            INSERT INTO notifications (title, message, type, date)
            VALUES (?, ?, 'warning', ?)
        """, (
            f"New {role.capitalize()} Signup Request: {name}",
            f"{name} ({user_id}) has applied for a {role.capitalize()} account. Review and approve in User Access.",
            datetime.date.today().isoformat()
        ))

        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "request_id": req_id,
            "message": f"Application submitted successfully! Your {role.capitalize()} account for '{name}' is awaiting Administrator approval. You will receive an email confirmation at {email} once approved."
        }), 201
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": f"Registration failed: {str(e)}"}), 500


@app.route("/api/signup-requests", methods=["GET"])
def get_signup_requests():
    """Admin only: List all signup applications."""
    status = request.args.get("status")
    conn = db.get_db_connection()
    if status:
        if status.lower() in ("declined", "rejected"):
            rows = conn.execute("SELECT * FROM signup_requests WHERE status IN ('Declined', 'Rejected') ORDER BY id DESC").fetchall()
        else:
            rows = conn.execute("SELECT * FROM signup_requests WHERE LOWER(status) = LOWER(?) ORDER BY id DESC", (status,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM signup_requests ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/signup-requests/<int:req_id>/approve", methods=["POST"])
def approve_signup_request(req_id):
    """Admin only: Approve signup request, activate faculty record in Faculty & Mentor Management, and dispatch welcome email."""
    conn = db.get_db_connection()
    req = conn.execute("SELECT * FROM signup_requests WHERE id = ?", (req_id,)).fetchone()

    if not req:
        conn.close()
        return jsonify({"success": False, "message": "Registration request not found."}), 404

    if req["status"] == "Approved":
        conn.close()
        return jsonify({"success": False, "message": "This application has already been approved."}), 400

    try:
        # 1. Create or Update User in `users` table with Active status and all submitted application details
        conn.execute("""
            INSERT OR REPLACE INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status, department, assigned_year, specialization)
            VALUES (?, ?, ?, ?, NULL, ?, ?, ?, ?, 'Active', ?, ?, ?)
        """, (
            req["user_id"],
            req["password"],
            req["role"],
            req["display_name"],
            req["subjects"],
            req["extra_roles"],
            req["email"],
            req["phone"],
            req["department"] if "department" in req.keys() and req["department"] else "CSE",
            req["assigned_year"] if "assigned_year" in req.keys() and req["assigned_year"] else "2nd Year",
            req["specialization"] if "specialization" in req.keys() and req["specialization"] else "Academic Counseling"
        ))

        # 2. Update request status to Approved
        conn.execute("UPDATE signup_requests SET status = 'Approved', reviewed_at = ? WHERE id = ?", (
            datetime.datetime.now().isoformat(), req_id
        ))

        # 3. Create formatted acceptance email
        subject = f"Welcome to EduStudent Sight — {req['role'].capitalize()} Account Activated"
        body_html = f"""
        <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;">
            <div style="background: #1e293b; color: #ffffff; padding: 24px; text-align: center;">
                <h2 style="margin: 0; font-size: 22px;">EduStudent Sight</h2>
                <p style="margin: 4px 0 0; color: #94a3b8; font-size: 13px;">Academic Intelligence Platform</p>
            </div>
            <div style="padding: 28px; background: #ffffff;">
                <h3 style="color: #0f172a; margin-top: 0;">Welcome, {req['display_name']}!</h3>
                <p style="color: #334155; line-height: 1.6;">Your registration application for the <strong>{req['role'].upper()}</strong> role has been approved by the Institutional Administrator.</p>
                <div style="background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 6px; padding: 18px; margin: 20px 0;">
                    <h4 style="margin: 0 0 12px; color: #0f172a; font-size: 14px; text-transform: uppercase;">Your Account Credentials:</h4>
                    <table style="width: 100%; font-size: 14px; color: #334155;">
                        <tr><td style="padding: 4px 0; width: 140px;"><strong>User ID:</strong></td><td><code style="background: #e2e8f0; padding: 2px 6px; border-radius: 4px;">{req['user_id']}</code></td></tr>
                        <tr><td style="padding: 4px 0;"><strong>Password:</strong></td><td><code style="background: #e2e8f0; padding: 2px 6px; border-radius: 4px;">{req['password']}</code></td></tr>
                        <tr><td style="padding: 4px 0;"><strong>Role:</strong></td><td><span style="color: #2563eb; font-weight: 600;">{req['role'].capitalize()}</span></td></tr>
                        <tr><td style="padding: 4px 0;"><strong>Assigned Subjects:</strong></td><td>{req['subjects'] or 'All Departmental Subjects'}</td></tr>
                        <tr><td style="padding: 4px 0;"><strong>Responsibilities:</strong></td><td>{req['extra_roles'] or 'Department Member'}</td></tr>
                    </table>
                </div>
                <p style="color: #334155; line-height: 1.6;">You may now log in to the platform with your credentials.</p>
            </div>
            <div style="background: #f1f5f9; padding: 16px; text-align: center; font-size: 12px; color: #64748b; border-top: 1px solid #e2e8f0;">
                &copy; 2026 EduStudent Sight • Vignan University Department of CSE
            </div>
        </div>
        """
        body_text = f"Welcome, {req['display_name']}! Your {req['role']} account ({req['user_id']}) has been activated. Password: {req['password']}."

        conn.commit()
        conn.close()

        # Log email
        _log_email(req["email"], subject, body_html, body_text, "Account Approved")

        return jsonify({
            "success": True,
            "message": f"Application for {req['display_name']} ({req['user_id']}) approved! Faculty record created in Faculty & Mentor Management. Credentials dispatched to {req['email']}."
        })
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": f"Approval failed: {str(e)}"}), 500


@app.route("/api/signup-requests/<int:req_id>/reject", methods=["POST"])
def reject_signup_request(req_id):
    """Admin only: Decline signup request with specified reasoning and dispatch decline notice."""
    data = request.get_json() or {}
    reason = data.get("reason", "").strip() or "Application details could not be verified by the Department Administrator."

    conn = db.get_db_connection()
    req = conn.execute("SELECT * FROM signup_requests WHERE id = ?", (req_id,)).fetchone()

    if not req:
        conn.close()
        return jsonify({"success": False, "message": "Registration request not found."}), 404

    try:
        # Update request status to Declined
        conn.execute("UPDATE signup_requests SET status = 'Declined', rejection_reason = ?, reviewed_at = ? WHERE id = ?", (
            reason, datetime.datetime.now().isoformat(), req_id
        ))

        # Ensure no active faculty account exists in users table
        conn.execute("DELETE FROM users WHERE LOWER(id) = LOWER(?) AND role IN ('faculty', 'mentor')", (req["user_id"],))

        # Create structured decline email
        subject = f"EduStudent Sight Application Status — {req['role'].capitalize()} Account"
        body_html = f"""
        <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: auto; border: 1px solid #e2e8f0; border-radius: 8px; overflow: hidden;">
            <div style="background: #1e293b; color: #ffffff; padding: 24px; text-align: center;">
                <h2 style="margin: 0; font-size: 22px;">EduStudent Sight</h2>
                <p style="margin: 4px 0 0; color: #94a3b8; font-size: 13px;">Academic Intelligence Platform</p>
            </div>
            <div style="padding: 28px; background: #ffffff;">
                <h3 style="color: #0f172a; margin-top: 0;">Application Status Notice</h3>
                <p style="color: #334155; line-height: 1.6;">Dear {req['display_name']},</p>
                <p style="color: #334155; line-height: 1.6;">Thank you for your interest in joining EduStudent Sight as a <strong>{req['role'].upper()}</strong>. After reviewing your registration application (ID: <code>{req['user_id']}</code>), the Department Administrator has decided to decline the request at this time.</p>
                <div style="background: #fef2f2; border-left: 4px solid #ef4444; padding: 16px; margin: 20px 0; border-radius: 0 6px 6px 0;">
                    <strong style="color: #991b1b; display: block; margin-bottom: 6px;">Reason for Decision:</strong>
                    <p style="color: #7f1d1d; margin: 0; font-size: 14px; line-height: 1.5;">{reason}</p>
                </div>
                <p style="color: #334155; line-height: 1.6;">If you believe this decision was made in error or if you have questions, please contact the Department Academic Advisory office.</p>
            </div>
            <div style="background: #f1f5f9; padding: 16px; text-align: center; font-size: 12px; color: #64748b; border-top: 1px solid #e2e8f0;">
                &copy; 2026 EduStudent Sight • Vignan University Department of CSE
            </div>
        </div>
        """
        body_text = f"Dear {req['display_name']}, your application for {req['role']} ({req['user_id']}) was declined. Reason: {reason}"

        conn.commit()
        conn.close()

        # Log email
        _log_email(req["email"], subject, body_html, body_text, "Account Declined")

        return jsonify({
            "success": True,
            "message": f"Application for {req['display_name']} ({req['user_id']}) marked as Declined. Reason recorded in history and email notice dispatched."
        })
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": f"Decline failed: {str(e)}"}), 500


@app.route("/api/emails", methods=["GET"])
def get_email_logs():
    """Admin only: Retrieve audit logs of all outgoing transactional emails."""
    conn = db.get_db_connection()
    logs = conn.execute("SELECT * FROM email_logs ORDER BY id DESC LIMIT 50").fetchall()
    conn.close()
    return jsonify([dict(r) for r in logs])


# =====================================================
# 1D. FACULTY MANAGEMENT (Admin → Faculty Module)
# Pure Management Area for Approved Faculty & Mentors
# =====================================================

@app.route("/api/faculty", methods=["GET"])
def get_faculty_list():
    """Admin: Returns approved faculty/mentor accounts for Faculty & Mentor Management, plus summary stats."""
    conn = db.get_db_connection()

    # 1. Get all approved faculty/mentor users
    users = conn.execute("""
        SELECT id, display_name, role, subjects, extra_roles, email, phone,
               COALESCE(status, 'Active') as status,
               COALESCE(department, 'CSE') as department,
               COALESCE(assigned_year, '2nd Year') as assigned_year,
               COALESCE(specialization, 'Academic Counseling') as specialization
        FROM users
        WHERE role IN ('faculty', 'mentor')
        ORDER BY display_name ASC
    """).fetchall()

    # 2. Get all signup requests for pending count and declined history
    requests = conn.execute("""
        SELECT id as req_id, user_id, display_name, role, subjects, extra_roles,
               email, phone, status, created_at, reviewed_at, rejection_reason,
               COALESCE(department, 'CSE') as department,
               COALESCE(assigned_year, '2nd Year') as assigned_year,
               COALESCE(specialization, 'Academic Counseling') as specialization
        FROM signup_requests
        ORDER BY id DESC
    """).fetchall()

    # 3. Get student counts per mentor (from interventions table)
    mentor_students = conn.execute("""
        SELECT mentor_id, COUNT(DISTINCT student_id) as student_count
        FROM interventions
        WHERE mentor_id IS NOT NULL
        GROUP BY mentor_id
    """).fetchall()
    mentor_student_map = {r["mentor_id"]: r["student_count"] for r in mentor_students}

    # 4. Get high-risk student counts per mentor
    mentor_risk = conn.execute("""
        SELECT i.mentor_id, COUNT(DISTINCT s.id) as high_risk_count
        FROM interventions i
        JOIN students s ON s.id = i.student_id
        WHERE i.mentor_id IS NOT NULL AND s.risk >= 60
        GROUP BY i.mentor_id
    """).fetchall()
    mentor_risk_map = {r["mentor_id"]: r["high_risk_count"] for r in mentor_risk}

    # 5. Total students in system
    total_students = conn.execute("SELECT COUNT(*) as cnt FROM students").fetchone()["cnt"]

    conn.close()

    # Build approved faculty list (ONLY approved accounts)
    approved_faculty = []
    for u in users:
        uid = u["id"]
        approved_faculty.append({
            "id": uid,
            "display_name": u["display_name"],
            "role": u["role"],
            "subjects": u["subjects"],
            "extra_roles": u["extra_roles"],
            "email": u["email"],
            "phone": u["phone"],
            "status": u["status"] or "Active",
            "department": u["department"],
            "assigned_year": u["assigned_year"],
            "specialization": u["specialization"],
            "source": "approved",
            "students_assigned": mentor_student_map.get(uid, 0),
            "high_risk_students": mentor_risk_map.get(uid, 0),
        })

    # Declined applications stored as history (not active accounts)
    declined_applications = []
    for r in requests:
        if r["status"] in ("Declined", "Rejected"):
            declined_applications.append({
                "id": r["user_id"],
                "req_id": r["req_id"],
                "display_name": r["display_name"],
                "role": r["role"],
                "subjects": r["subjects"],
                "extra_roles": r["extra_roles"],
                "email": r["email"],
                "phone": r["phone"],
                "status": "Declined",
                "department": r["department"],
                "assigned_year": r["assigned_year"],
                "specialization": r["specialization"],
                "source": "declined",
                "created_at": r["created_at"],
                "reviewed_at": r["reviewed_at"],
                "rejection_reason": r["rejection_reason"],
                "students_assigned": 0,
                "high_risk_students": 0,
            })

    pending_count = len([r for r in requests if r["status"] == "Pending"])

    # Summary stats
    summary = {
        "total_faculty": len(approved_faculty),
        "active_faculty": len([f for f in approved_faculty if f["status"] == "Active"]),
        "pending_applications": pending_count,
        "mentors": len([f for f in approved_faculty if f["role"] == "mentor"]),
        "total_students_assigned": sum(f["students_assigned"] for f in approved_faculty),
        "total_students": total_students,
    }

    return jsonify({
        "faculty": approved_faculty,
        "declined_history": declined_applications,
        "summary": summary
    })


@app.route("/api/mentors", methods=["GET"])
def get_mentors_list():
    """Returns active mentors with department, assigned year cohort, and current mentees count."""
    conn = db.get_db_connection()
    dept = request.args.get("department")
    year = request.args.get("year")

    query = """
        SELECT u.id, u.display_name, u.role, u.email, u.phone, u.subjects, u.extra_roles,
               COALESCE(u.status, 'Active') as status,
               COALESCE(u.department, 'CSE') as department,
               COALESCE(u.assigned_year, '2nd Year') as assigned_year,
               COALESCE(u.specialization, 'Academic Counseling') as specialization,
               COUNT(DISTINCT i.student_id) as active_mentees
        FROM users u
        LEFT JOIN interventions i ON LOWER(u.id) = LOWER(i.mentor_id) AND i.status != 'Completed'
        WHERE u.role IN ('mentor', 'faculty') AND COALESCE(u.status, 'Active') = 'Active'
    """
    params = []
    if dept and dept != "ALL":
        query += " AND (LOWER(u.department) = LOWER(?) OR LOWER(u.subjects) LIKE LOWER(?))"
        params.extend([dept, f"%{dept}%"])
    if year and year != "ALL":
        query += " AND (LOWER(u.assigned_year) = LOWER(?) OR u.assigned_year = 'All Years')"
        params.append(year)

    query += " GROUP BY u.id ORDER BY u.role DESC, u.display_name ASC"
    rows = conn.execute(query, tuple(params)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/faculty/<faculty_id>", methods=["GET"])
def get_faculty_detail(faculty_id):
    """Admin: Full faculty detail — profile, application, students, interventions, AI risks."""
    conn = db.get_db_connection()

    # 1. Get user account (if approved)
    user = conn.execute("""
        SELECT id, display_name, role, subjects, extra_roles, email, phone,
               COALESCE(status, 'Active') as status,
               COALESCE(department, 'CSE') as department,
               COALESCE(assigned_year, '2nd Year') as assigned_year,
               COALESCE(specialization, 'Academic Counseling') as specialization
        FROM users WHERE LOWER(id) = LOWER(?)
    """, (faculty_id,)).fetchone()

    # 2. Get original signup application (if exists)
    application = conn.execute("""
        SELECT * FROM signup_requests
        WHERE LOWER(user_id) = LOWER(?)
        ORDER BY id DESC LIMIT 1
    """, (faculty_id,)).fetchone()

    if not user and not application:
        conn.close()
        return jsonify({"success": False, "message": "Faculty not found."}), 404

    # Build profile from whichever source exists
    profile = {}
    if user:
        profile = {
            "id": user["id"],
            "display_name": user["display_name"],
            "role": user["role"],
            "subjects": user["subjects"],
            "extra_roles": user["extra_roles"],
            "email": user["email"],
            "phone": user["phone"],
            "status": user["status"] or "Active",
            "department": user["department"],
            "assigned_year": user["assigned_year"],
            "specialization": user["specialization"],
            "source": "approved",
        }
    elif application:
        profile = {
            "id": application["user_id"],
            "display_name": application["display_name"],
            "role": application["role"],
            "subjects": application["subjects"],
            "extra_roles": application["extra_roles"],
            "email": application["email"],
            "phone": application["phone"],
            "status": application["status"],
            "department": application["department"] if "department" in application.keys() else "CSE",
            "assigned_year": application["assigned_year"] if "assigned_year" in application.keys() else "2nd Year",
            "specialization": application["specialization"] if "specialization" in application.keys() else "Academic Counseling",
            "source": "application",
        }

    # Application details
    app_details = None
    if application:
        app_details = {
            "req_id": application["id"],
            "status": application["status"],
            "submitted_date": application["created_at"],
            "reviewed_date": application["reviewed_at"],
            "rejection_reason": application["rejection_reason"],
        }

    # 3. Assigned students (from interventions)
    assigned_students_rows = conn.execute("""
        SELECT DISTINCT s.id, s.name, s.course, s.year, s.attendance, s.cgpa, s.risk
        FROM interventions i
        JOIN students s ON s.id = i.student_id
        WHERE LOWER(i.mentor_id) = LOWER(?)
        ORDER BY s.risk DESC
    """, (faculty_id,)).fetchall()
    assigned_students = [dict(s) for s in assigned_students_rows]

    # 4. Interventions by this faculty
    interventions = conn.execute("""
        SELECT id, student_id, date, action, status, notes, urgency
        FROM interventions
        WHERE LOWER(mentor_id) = LOWER(?)
        ORDER BY id DESC LIMIT 10
    """, (faculty_id,)).fetchall()

    # 5. AI risk summary — high-risk students in their purview
    high_risk = [s for s in assigned_students if s.get("risk", 0) >= 60]
    medium_risk = [s for s in assigned_students if 30 <= s.get("risk", 0) < 60]

    conn.close()

    return jsonify({
        "success": True,
        "profile": profile,
        "application": app_details,
        "assigned_students": assigned_students,
        "interventions": [dict(i) for i in interventions],
        "ai_summary": {
            "high_risk_count": len(high_risk),
            "medium_risk_count": len(medium_risk),
            "total_assigned": len(assigned_students),
            "pending_interventions": len([i for i in interventions if dict(i).get("status") == "Pending"]),
            "completed_interventions": len([i for i in interventions if dict(i).get("status") == "Completed"]),
        }
    })


@app.route("/api/faculty/<faculty_id>", methods=["PUT"])
def update_faculty(faculty_id):
    """Admin: Edit faculty profile information."""
    data = request.get_json() or {}
    conn = db.get_db_connection()

    user = conn.execute("SELECT id FROM users WHERE LOWER(id) = LOWER(?)", (faculty_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"success": False, "message": "Faculty account not found. Only approved faculty can be edited."}), 404

    # Update allowed fields
    conn.execute("""
        UPDATE users SET
            display_name = COALESCE(?, display_name),
            subjects = COALESCE(?, subjects),
            extra_roles = COALESCE(?, extra_roles),
            email = COALESCE(?, email),
            phone = COALESCE(?, phone),
            department = COALESCE(?, department),
            assigned_year = COALESCE(?, assigned_year),
            specialization = COALESCE(?, specialization)
        WHERE LOWER(id) = LOWER(?)
    """, (
        data.get("display_name"),
        data.get("subjects"),
        data.get("extra_roles"),
        data.get("email"),
        data.get("phone"),
        data.get("department"),
        data.get("assigned_year"),
        data.get("specialization"),
        faculty_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Faculty {faculty_id} profile updated."})


@app.route("/api/faculty/<faculty_id>/status", methods=["PUT"])
def update_faculty_status(faculty_id):
    """Admin: Change faculty status (Active/Inactive)."""
    data = request.get_json() or {}
    new_status = data.get("status", "Active")

    if new_status not in ("Active", "Inactive"):
        return jsonify({"success": False, "message": "Status must be 'Active' or 'Inactive'."}), 400

    conn = db.get_db_connection()
    user = conn.execute("SELECT id FROM users WHERE LOWER(id) = LOWER(?)", (faculty_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"success": False, "message": "Faculty account not found."}), 404

    conn.execute("UPDATE users SET status = ? WHERE LOWER(id) = LOWER(?)", (new_status, faculty_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Faculty {faculty_id} status changed to {new_status}."})


@app.route("/api/faculty/bulk-import", methods=["POST"])
def bulk_import_faculty():
    """Admin: Bulk import faculty from JSON array. Validates each row before inserting."""
    data = request.get_json() or {}
    rows = data.get("rows", [])
    if not rows:
        return jsonify({"success": False, "message": "No rows provided."}), 400

    conn = db.get_db_connection()

    # Pre-fetch existing IDs and emails for duplicate checking
    existing_users = conn.execute("SELECT LOWER(id) as id, LOWER(COALESCE(email,'')) as email FROM users").fetchall()
    existing_ids = set(r["id"] for r in existing_users)
    existing_emails = set(r["email"] for r in existing_users if r["email"])

    existing_reqs = conn.execute("SELECT LOWER(user_id) as uid, LOWER(COALESCE(email,'')) as email FROM signup_requests WHERE status = 'Pending'").fetchall()
    pending_ids = set(r["uid"] for r in existing_reqs)
    pending_emails = set(r["email"] for r in existing_reqs if r["email"])

    results = []
    imported = 0
    failed = 0
    duplicates = 0

    # Track IDs and emails within the import batch itself
    batch_ids = set()
    batch_emails = set()

    for idx, row in enumerate(rows):
        user_id = str(row.get("id", row.get("user_id", ""))).strip()
        name = str(row.get("display_name", row.get("name", row.get("full_name", "")))).strip()
        email = str(row.get("email", "")).strip()
        phone = str(row.get("phone", "")).strip()
        role = str(row.get("role", "faculty")).strip().lower()
        subjects = str(row.get("subjects", row.get("assigned_subjects", ""))).strip()
        extra_roles = str(row.get("extra_roles", row.get("additional_responsibilities", ""))).strip()
        password = str(row.get("password", user_id)).strip()

        errors = []

        # Required field validation
        if not user_id:
            errors.append("Missing User ID")
        if not name:
            errors.append("Missing Full Name")
        if not email:
            errors.append("Missing Email")

        # Email format validation
        if email and ("@" not in email or "." not in email.split("@")[-1]):
            errors.append("Invalid email format")

        # Role validation
        if role not in ("faculty", "mentor"):
            errors.append(f"Invalid role '{role}' — must be 'faculty' or 'mentor'")

        # Duplicate checks — existing database
        if user_id and user_id.lower() in existing_ids:
            errors.append("Duplicate — User ID already exists in system")
            duplicates += 1
        elif user_id and user_id.lower() in pending_ids:
            errors.append("Duplicate — User ID has a pending application")
            duplicates += 1

        if email and email.lower() in existing_emails:
            errors.append("Duplicate — Email already registered")
            if "Duplicate" not in errors[0] if errors else "":
                duplicates += 1
        elif email and email.lower() in pending_emails:
            errors.append("Duplicate — Email has a pending application")

        # Batch-internal duplicate checks
        if user_id and user_id.lower() in batch_ids:
            errors.append("Duplicate — User ID appears multiple times in this import")
        if email and email.lower() in batch_emails:
            errors.append("Duplicate — Email appears multiple times in this import")

        row_result = {
            "row": idx + 1,
            "id": user_id,
            "name": name,
            "email": email,
            "role": role,
            "status": "valid" if not errors else ("duplicate" if any("Duplicate" in e for e in errors) else "invalid"),
            "errors": errors
        }

        if not errors:
            try:
                conn.execute("""
                    INSERT INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone, status)
                    VALUES (?, ?, ?, ?, NULL, ?, ?, ?, ?, 'Active')
                """, (user_id, password, role, name, subjects, extra_roles, email, phone))
                imported += 1
                row_result["status"] = "imported"
                # Track in batch sets
                batch_ids.add(user_id.lower())
                if email:
                    batch_emails.add(email.lower())
                # Also add to existing sets so subsequent rows catch intra-batch dupes
                existing_ids.add(user_id.lower())
                if email:
                    existing_emails.add(email.lower())
            except Exception as e:
                row_result["status"] = "error"
                row_result["errors"].append(str(e))
                failed += 1
        else:
            failed += 1

        results.append(row_result)

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "imported": imported,
        "failed": failed,
        "duplicates": duplicates,
        "total": len(rows),
        "results": results,
        "message": f"Import complete: {imported} imported, {failed} failed, {duplicates} duplicates."
    })


@app.route("/api/faculty/bulk-delete", methods=["POST"])
def bulk_delete_faculty():
    """Admin: Delete or deactivate multiple faculty. Checks dependencies first."""
    data = request.get_json() or {}
    ids = data.get("ids", [])
    force = data.get("force", False)

    if not ids:
        return jsonify({"success": False, "message": "No faculty IDs provided."}), 400

    conn = db.get_db_connection()

    deleted = 0
    deactivated = 0
    skipped = []
    details = []

    for fid in ids:
        # Check if user exists
        user = conn.execute("SELECT id, display_name, role FROM users WHERE LOWER(id) = LOWER(?)", (fid,)).fetchone()
        if not user:
            skipped.append({"id": fid, "reason": "User not found"})
            continue

        # Check for dependencies: interventions, signup requests
        intervention_count = conn.execute("SELECT COUNT(*) as cnt FROM interventions WHERE LOWER(mentor_id) = LOWER(?)", (fid,)).fetchone()["cnt"]
        signup_count = conn.execute("SELECT COUNT(*) as cnt FROM signup_requests WHERE LOWER(user_id) = LOWER(?)", (fid,)).fetchone()["cnt"]
        has_dependencies = intervention_count > 0 or signup_count > 0

        if has_dependencies and not force:
            # Deactivate instead of delete
            conn.execute("UPDATE users SET status = 'Inactive' WHERE LOWER(id) = LOWER(?)", (fid,))
            deactivated += 1
            details.append({
                "id": fid,
                "name": user["display_name"],
                "action": "deactivated",
                "reason": f"Has {intervention_count} interventions, {signup_count} application records"
            })
        else:
            # Safe to delete — remove user record only (preserve signup_requests for history)
            conn.execute("DELETE FROM users WHERE LOWER(id) = LOWER(?)", (fid,))
            deleted += 1
            details.append({
                "id": fid,
                "name": user["display_name"],
                "action": "deleted",
                "reason": "No dependencies" if not has_dependencies else "Force deleted"
            })

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "deleted": deleted,
        "deactivated": deactivated,
        "skipped": skipped,
        "details": details,
        "message": f"Processed {len(ids)} faculty: {deleted} deleted, {deactivated} deactivated, {len(skipped)} skipped."
    })


@app.route("/api/faculty/bulk-status", methods=["POST"])
def bulk_status_faculty():
    """Admin: Change status of multiple faculty at once."""
    data = request.get_json() or {}
    ids = data.get("ids", [])
    new_status = data.get("status", "Active")

    if not ids:
        return jsonify({"success": False, "message": "No faculty IDs provided."}), 400
    if new_status not in ("Active", "Inactive"):
        return jsonify({"success": False, "message": "Status must be 'Active' or 'Inactive'."}), 400

    conn = db.get_db_connection()
    updated = 0
    for fid in ids:
        result = conn.execute("UPDATE users SET status = ? WHERE LOWER(id) = LOWER(?) AND role IN ('faculty', 'mentor')", (new_status, fid))
        if result.rowcount > 0:
            updated += 1
    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "updated": updated,
        "message": f"{updated} faculty members set to {new_status}."
    })


# =====================================================
# 2. STUDENT CRUD
# =====================================================

def _get_thresholds():
    """Helper: Read dynamic thresholds from settings."""
    settings = db.get_system_settings()
    return {
        "attendance_threshold": float(settings.get("attendance_threshold", 75)),
        "risk_cgpa_threshold": float(settings.get("risk_cgpa_threshold", 7.5)),
        "lms_threshold": float(settings.get("lms_threshold", 60)),
        "assignment_threshold": float(settings.get("assignment_threshold", 70)),
        "safe_risk_threshold": float(settings.get("safe_risk_threshold", 30)),
        "high_risk_threshold": float(settings.get("high_risk_threshold", 65)),
        "critical_risk_threshold": float(settings.get("critical_risk_threshold", 80))
    }


@app.route("/api/students", methods=["GET"])
def get_students():
    conn = db.get_db_connection()
    rows = conn.execute("SELECT * FROM students ORDER BY risk DESC").fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])


@app.route("/api/students/<student_id>", methods=["GET"])
def get_student_detail(student_id):
    conn = db.get_db_connection()
    row = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
    interventions = conn.execute("""
        SELECT i.*, u.display_name as mentor_name, u.email as mentor_email
        FROM interventions i
        LEFT JOIN users u ON LOWER(i.mentor_id) = LOWER(u.id)
        WHERE i.student_id = ?
        ORDER BY i.id DESC
    """, (student_id,)).fetchall()
    marks = conn.execute("""
        SELECT sm.*, s.name as subject_name, s.short_name
        FROM subject_marks sm
        LEFT JOIN subjects s ON sm.subject_code = s.code
        WHERE sm.student_id = ?
    """, (student_id,)).fetchall()
    activities = conn.execute("""
        SELECT sa.*, e.name as activity_name, e.category
        FROM student_activities sa
        JOIN extracurriculars e ON sa.activity_id = e.id
        WHERE sa.student_id = ?
    """, (student_id,)).fetchall()
    conn.close()

    if not row:
        return jsonify({"error": "Student not found"}), 404

    thresholds = _get_thresholds()
    student_dict = dict(row)
    student_dict["interventions"] = [dict(i) for i in interventions]
    student_dict["subject_marks"] = [dict(m) for m in marks]
    student_dict["activities"] = [dict(a) for a in activities]
    student_dict["ai_analysis"] = ai_engine.calculate_risk_score(
        student_dict.get("attendance"),
        student_dict.get("cgpa"),
        student_dict.get("lms_score"),
        **thresholds
    )
    return jsonify(student_dict)


@app.route("/api/students", methods=["POST"])
def add_student():
    data = request.get_json() or {}
    student_id = data.get("id", "").strip()
    name = data.get("name", "").strip()
    if not student_id or not name:
        return jsonify({"success": False, "message": "Student ID and Name are required."}), 400

    attendance = int(data.get("attendance", 0))
    cgpa = float(data.get("cgpa", 0.0))
    lms_score = int(data.get("lms_score", attendance))
    email = data.get("email", "").strip() or "Unknown"
    phone = data.get("phone", "").strip() or "Unknown"

    thresholds = _get_thresholds()
    risk_info = ai_engine.calculate_risk_score(attendance, cgpa, lms_score, **thresholds)
    risk_score = risk_info["risk_score"]

    conn = db.get_db_connection()
    try:
        conn.execute("""
            INSERT INTO students (id, name, gender, course, year, cgpa, credits, attendance, lms_score, risk, father, mother, mother_tongue, place, region, country, email, phone)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            student_id, name, data.get("gender", "Unknown"), data.get("course", "Unknown"),
            data.get("year", "Unknown"), cgpa, int(data.get("credits", 0)), attendance,
            lms_score, risk_score, data.get("father", "Unknown"), data.get("mother", "Unknown"),
            data.get("motherTongue", "Unknown"), data.get("place", "Unknown"),
            data.get("region", "Unknown"), data.get("country", "Unknown"),
            email, phone
        ))
        # Also create a user account for the student
        conn.execute("INSERT OR IGNORE INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone) VALUES (?,?,?,?,?,?,?,?,?)", (
            student_id, student_id, "student", name, student_id, None, None, email, phone
        ))
        conn.commit()
        conn.close()
        return jsonify({"success": True, "student": {**data, "risk": risk_score}}), 201
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": f"Error: {str(e)}"}), 500


@app.route("/api/students/<student_id>", methods=["PUT"])
def update_student(student_id):
    data = request.get_json() or {}
    attendance = int(data.get("attendance", 80))
    cgpa = float(data.get("cgpa", 7.5))
    lms_score = int(data.get("lms_score", attendance))

    thresholds = _get_thresholds()
    risk_info = ai_engine.calculate_risk_score(attendance, cgpa, lms_score, **thresholds)
    risk_score = risk_info["risk_score"]

    conn = db.get_db_connection()
    conn.execute("""
        UPDATE students SET name=?, course=?, year=?, cgpa=?, credits=?, attendance=?, lms_score=?, risk=?, father=?, mother=?, mother_tongue=?, place=?, region=?, email=?, phone=?
        WHERE id=?
    """, (
        data.get("name"), data.get("course"), data.get("year"), cgpa,
        int(data.get("credits", 24)), attendance, lms_score, risk_score,
        data.get("father"), data.get("mother"), data.get("motherTongue"),
        data.get("place"), data.get("region"), data.get("email"), data.get("phone"),
        student_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Student updated.", "risk": risk_score})


@app.route("/api/students/<student_id>", methods=["DELETE"])
def delete_student(student_id):
    conn = db.get_db_connection()
    conn.execute("DELETE FROM interventions WHERE student_id = ?", (student_id,))
    conn.execute("DELETE FROM subject_marks WHERE student_id = ?", (student_id,))
    conn.execute("DELETE FROM student_activities WHERE student_id = ?", (student_id,))
    conn.execute("DELETE FROM users WHERE id = ? AND role = 'student'", (student_id,))
    conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Student {student_id} deleted."})


@app.route("/api/students/bulk-delete", methods=["POST"])
def bulk_delete_students():
    """Admin/Faculty: Delete multiple students at once."""
    data = request.get_json() or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"success": False, "message": "No student IDs provided."}), 400

    conn = db.get_db_connection()
    deleted = 0
    for sid in ids:
        conn.execute("DELETE FROM interventions WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM subject_marks WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM student_activities WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM notifications WHERE student_id = ?", (sid,))
        conn.execute("DELETE FROM users WHERE id = ? AND role = 'student'", (sid,))
        conn.execute("DELETE FROM students WHERE id = ?", (sid,))
        deleted += 1
    conn.commit()
    conn.close()
    return jsonify({"success": True, "deleted": deleted, "message": f"Deleted {deleted} student records."})


@app.route("/api/recalculate-risks", methods=["POST"])
def recalculate_risks():
    """Recalculates risk scores for ALL students using current threshold settings."""
    thresholds = _get_thresholds()
    conn = db.get_db_connection()
    rows = conn.execute("SELECT id, attendance, cgpa, lms_score FROM students").fetchall()
    count = 0
    for r in rows:
        risk_info = ai_engine.calculate_risk_score(r["attendance"], r["cgpa"], r["lms_score"], **thresholds)
        conn.execute("UPDATE students SET risk = ? WHERE id = ?", (risk_info["risk_score"], r["id"]))
        count += 1
    conn.commit()
    conn.close()
    return jsonify({"success": True, "updated_count": count})


# =====================================================
# 2.1 24-HOUR AUTOMATED ATTENDANCE PROGRESSION ENGINE (+2%)
# =====================================================

def _run_daily_attendance_increment(force=False, increment_pct=2.0):
    conn = db.get_db_connection()
    c = conn.cursor()
    
    # Read last progression time
    row = c.execute("SELECT value FROM settings WHERE key = 'last_attendance_progression'").fetchone()
    last_progression_str = row["value"] if row else None
    
    now = datetime.datetime.now()
    should_run = force
    
    if not force:
        if not last_progression_str:
            should_run = True
        else:
            try:
                last_dt = datetime.datetime.fromisoformat(last_progression_str)
                if (now - last_dt).total_seconds() >= 86400:  # 24 hours
                    should_run = True
            except Exception:
                should_run = True

    if not should_run:
        conn.close()
        return {
            "applied": False,
            "reason": "Less than 24 hours since last progression",
            "last_progression": last_progression_str,
            "next_in_seconds": max(0, 86400 - int((now - datetime.datetime.fromisoformat(last_progression_str)).total_seconds())) if last_progression_str else 0
        }

    thresholds = _get_thresholds()
    students_rows = c.execute("SELECT id, attendance, cgpa, lms_score FROM students").fetchall()
    updated_count = 0
    
    for s in students_rows:
        curr_att = float(s["attendance"] or 0)
        new_att = min(100.0, round(curr_att + increment_pct, 1))
        risk_info = ai_engine.calculate_risk_score(new_att, s["cgpa"], s["lms_score"], **thresholds)
        c.execute("UPDATE students SET attendance = ?, risk = ? WHERE id = ?", (int(new_att), risk_info["risk_score"], s["id"]))
        c.execute("UPDATE subject_marks SET attendance = MIN(100, attendance + ?) WHERE student_id = ?", (int(increment_pct), s["id"]))
        updated_count += 1

    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('last_attendance_progression', ?)", (now.isoformat(),))
    
    # Audit log notification
    c.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES (?, ?, 'info', ?, NULL, 0)
    """, (
        f"Daily Attendance Progression (+{increment_pct}%)",
        f"24-Hour Cycle: Attendance increased by +{increment_pct}% for {updated_count} students. AI risk scores recalibrated.",
        now.strftime("%b %d, %Y %I:%M %p")
    ))
    
    conn.commit()
    conn.close()
    
    return {
        "applied": True,
        "students_updated": updated_count,
        "increment_pct": increment_pct,
        "timestamp": now.isoformat(),
        "message": f"Successfully increased attendance by +{increment_pct}% for {updated_count} students."
    }


@app.route("/api/attendance/advance-day", methods=["POST"])
def advance_attendance_day():
    """Manual/Automated trigger: Advance calendar by 1 day (+24h) and increase attendance by +2%."""
    data = request.get_json() or {}
    increment_pct = float(data.get("increment_pct", 2.0))
    res = _run_daily_attendance_increment(force=True, increment_pct=increment_pct)
    return jsonify({"success": True, **res})


@app.route("/api/attendance/status", methods=["GET"])
def get_attendance_progression_status():
    """Returns the last 24h attendance progression timestamp and next scheduled trigger."""
    conn = db.get_db_connection()
    row = conn.execute("SELECT value FROM settings WHERE key = 'last_attendance_progression'").fetchone()
    conn.close()
    
    last_str = row["value"] if row else None
    now = datetime.datetime.now()
    hours_since = None
    next_in_hours = 0
    
    if last_str:
        try:
            last_dt = datetime.datetime.fromisoformat(last_str)
            elapsed_sec = (now - last_dt).total_seconds()
            hours_since = round(elapsed_sec / 3600.0, 1)
            next_in_hours = max(0.0, round((86400 - elapsed_sec) / 3600.0, 1))
        except Exception:
            pass

    return jsonify({
        "success": True,
        "last_progression": last_str,
        "hours_since_last": hours_since,
        "next_in_hours": next_in_hours,
        "increment_rate": "+2.0% per 24 hours"
    })



# =====================================================
# 3. SUBJECTS & MARKS
# =====================================================

@app.route("/api/subjects", methods=["GET"])
def get_subjects():
    year = request.args.get("year")
    conn = db.get_db_connection()
    if year:
        rows = conn.execute("SELECT * FROM subjects WHERE year = ? ORDER BY semester, code", (year,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM subjects ORDER BY year, semester, code").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/subject-marks/<student_id>", methods=["GET"])
def get_subject_marks(student_id):
    conn = db.get_db_connection()
    rows = conn.execute("""
        SELECT sm.*, s.name as subject_name, s.short_name, s.credits, s.year, s.semester
        FROM subject_marks sm
        JOIN subjects s ON sm.subject_code = s.code
        WHERE sm.student_id = ?
        ORDER BY s.semester, s.code
    """, (student_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/subject-marks", methods=["GET"])
def get_all_subject_marks():
    conn = db.get_db_connection()
    rows = conn.execute("""
        SELECT sm.*, s.name as subject_name, s.short_name
        FROM subject_marks sm
        JOIN subjects s ON sm.subject_code = s.code
        ORDER BY sm.student_id, s.code
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/subject-marks", methods=["POST"])
def upsert_subject_marks():
    data = request.get_json() or {}
    conn = db.get_db_connection()
    conn.execute("""
        INSERT INTO subject_marks (student_id, subject_code, attendance, internal_marks, external_marks, assignment_score, grade)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(student_id, subject_code) DO UPDATE SET
            attendance=excluded.attendance, internal_marks=excluded.internal_marks,
            external_marks=excluded.external_marks, assignment_score=excluded.assignment_score, grade=excluded.grade
    """, (
        data.get("student_id"), data.get("subject_code"),
        int(data.get("attendance", 0)), int(data.get("internal_marks", 0)),
        int(data.get("external_marks", 0)), int(data.get("assignment_score", 0)),
        data.get("grade", "N/A")
    ))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Subject marks saved."})


# =====================================================
# 4. EXTRACURRICULAR ACTIVITIES
# =====================================================

@app.route("/api/extracurriculars", methods=["GET"])
def get_extracurriculars():
    conn = db.get_db_connection()
    rows = conn.execute("SELECT * FROM extracurriculars ORDER BY category, name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/student-activities/<student_id>", methods=["GET"])
def get_student_activities(student_id):
    conn = db.get_db_connection()
    rows = conn.execute("""
        SELECT sa.*, e.name as activity_name, e.category, e.description as activity_desc
        FROM student_activities sa
        JOIN extracurriculars e ON sa.activity_id = e.id
        WHERE sa.student_id = ?
        ORDER BY e.category, e.name
    """, (student_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# =====================================================
# 5. INTERVENTIONS (Full Lifecycle)
# =====================================================

@app.route("/api/interventions", methods=["GET"])
def get_all_interventions():
    mentor_id = request.args.get("mentor_id")
    student_id = request.args.get("student_id")
    caller_role = request.args.get("role", "").lower()
    caller_id = request.args.get("user_id", "")
    conn = db.get_db_connection()
    
    query = """
        SELECT i.*, 
               s.name as student_name, s.course as student_course, s.year as student_year, s.risk as student_risk,
               s.attendance as student_attendance, s.cgpa as student_cgpa,
               u.display_name as mentor_name, u.email as mentor_email, 
               COALESCE(u.department, 'CSE') as mentor_dept, 
               COALESCE(u.assigned_year, '2nd Year') as mentor_year,
               creator.display_name as creator_name
        FROM interventions i
        LEFT JOIN students s ON i.student_id = s.id
        LEFT JOIN users u ON LOWER(i.mentor_id) = LOWER(u.id)
        LEFT JOIN users creator ON LOWER(i.created_by) = LOWER(creator.id)
    """
    params = []
    conditions = []
    
    if mentor_id and mentor_id != "ALL":
        conditions.append("(LOWER(i.mentor_id) = LOWER(?) OR LOWER(i.created_by) = LOWER(?))")
        params.extend([mentor_id, mentor_id])
    elif caller_role == "mentor" and caller_id:
        conditions.append("(LOWER(i.mentor_id) = LOWER(?) OR LOWER(i.created_by) = LOWER(?))")
        params.extend([caller_id, caller_id])

    if student_id:
        conditions.append("i.student_id = ?")
        params.append(student_id)
        
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY i.id DESC"
    
    rows = conn.execute(query, tuple(params)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/interventions/<student_id>", methods=["GET"])
def get_student_interventions(student_id):
    conn = db.get_db_connection()
    rows = conn.execute("""
        SELECT i.*, 
               s.name as student_name, s.course as student_course, s.year as student_year, s.risk as student_risk,
               s.attendance as student_attendance, s.cgpa as student_cgpa,
               u.display_name as mentor_name, u.email as mentor_email,
               COALESCE(u.department, 'CSE') as mentor_dept, 
               COALESCE(u.assigned_year, '2nd Year') as mentor_year,
               creator.display_name as creator_name
        FROM interventions i
        LEFT JOIN students s ON i.student_id = s.id
        LEFT JOIN users u ON LOWER(i.mentor_id) = LOWER(u.id)
        LEFT JOIN users creator ON LOWER(i.created_by) = LOWER(creator.id)
        WHERE i.student_id = ?
        ORDER BY i.id DESC
    """, (student_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/interventions", methods=["POST"])
def create_intervention():
    data = request.get_json() or {}
    student_id = data.get("student_id")
    mentor_id = data.get("mentor_id")
    created_by = data.get("created_by")
    action = data.get("action", "1-on-1 Academic Counseling")
    urgency = data.get("urgency", "Moderate")
    session_date = data.get("date") or datetime.date.today().isoformat()
    session_time = data.get("session_time", "10:00 AM")
    location = data.get("location", "Mentorship Cabin 204")
    notes = data.get("notes", "")
    subject_code = data.get("subject_code", "General")

    conn = db.get_db_connection()

    # Look up student details
    student = conn.execute("SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()

    # Auto-resolve mentor if needed
    if not mentor_id or mentor_id == "auto":
        st_course = student["course"] if student else "CSE"
        st_year = student["year"] if student else "2nd Year"
        matched_mentor = conn.execute("""
            SELECT id FROM users
            WHERE role = 'mentor' AND (LOWER(department) = LOWER(?) OR department = 'ALL')
              AND (LOWER(assigned_year) = LOWER(?) OR assigned_year = 'All Years')
            LIMIT 1
        """, (st_course, st_year)).fetchone()
        
        if matched_mentor:
            mentor_id = matched_mentor["id"]
        else:
            fallback = conn.execute("SELECT id FROM users WHERE role = 'mentor' LIMIT 1").fetchone()
            mentor_id = fallback["id"] if fallback else (created_by or "MEN001")

    # Insert intervention record
    c = conn.cursor()
    c.execute("""
        INSERT INTO interventions (student_id, date, action, status, notes, urgency, subject_code, mentor_id, created_by, session_time, location)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        student_id,
        session_date,
        action,
        "In Progress" if data.get("status") in ("In Progress", "Pending", "Scheduled") else (data.get("status") or "In Progress"),
        notes,
        urgency,
        subject_code,
        mentor_id,
        created_by,
        session_time,
        location
    ))
    int_id = c.lastrowid

    # Get mentor name for notifications
    mentor_user = conn.execute("SELECT display_name FROM users WHERE LOWER(id) = LOWER(?)", (mentor_id,)).fetchone()
    mentor_name = mentor_user["display_name"] if mentor_user else mentor_id
    st_name = student["name"] if student else student_id

    # Dispatch notification to student
    now_str = datetime.datetime.now().strftime("%b %d, %Y %I:%M %p")
    notif_type = "danger" if urgency == "Critical" else "info"
    c.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES (?, ?, ?, ?, ?, 0)
    """, (
        f"📅 New Session: {action}",
        f"1-on-1 session scheduled with {mentor_name} on {session_date} at {session_time} ({location}). Notes: {notes or 'Routine academic review'}",
        notif_type,
        now_str,
        student_id
    ))

    conn.commit()
    conn.close()
    return jsonify({
        "success": True,
        "message": f"Session scheduled successfully for {st_name} with {mentor_name}.",
        "intervention_id": int_id,
        "assigned_mentor": mentor_id
    })


@app.route("/api/interventions/update/<int:intervention_id>", methods=["PUT", "POST"])
def update_intervention(intervention_id):
    """Direct status update. Admin & Faculty can mark completed; Mentor can mark completed if they are the initiator."""
    data = request.get_json() or {}
    new_status = data.get("status", "In Progress")
    caller_role = (data.get("caller_role") or "").lower()
    caller_id = (data.get("user_id") or "").lower()

    conn = db.get_db_connection()
    intervention = conn.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    if not intervention:
        conn.close()
        return jsonify({"success": False, "message": "Intervention not found."}), 404

    # UAC Check for direct Completion:
    if new_status == "Completed":
        is_admin_or_faculty = caller_role in ("admin", "faculty")
        created_by = (intervention["created_by"] or "").lower()
        is_initiator = not created_by or created_by == caller_id

        if not is_admin_or_faculty and not is_initiator:
            conn.close()
            return jsonify({
                "success": False, 
                "message": "This session was scheduled by another staff member. Please use 'Request Completion Review' so the initiator can verify and approve."
            }), 403

    completed_date = None
    if new_status == "Completed":
        completed_date = datetime.date.today().isoformat()

    conn.execute("""
        UPDATE interventions SET 
            status = ?, 
            notes = COALESCE(?, notes), 
            completed_date = ?,
            reviewed_by = COALESCE(?, reviewed_by),
            reviewed_at = CASE WHEN ? = 'Completed' THEN ? ELSE reviewed_at END
        WHERE id = ?
    """, (new_status, data.get("notes"), completed_date, data.get("user_id"), new_status, datetime.datetime.now().isoformat(), intervention_id))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Intervention #{intervention_id} updated to {new_status}."})


@app.route("/api/interventions/<int:intervention_id>/request-completion", methods=["POST"])
def request_intervention_completion(intervention_id):
    """Mentor requests completion review for an intervention assigned by admin or faculty."""
    data = request.get_json() or {}
    mentor_id = data.get("mentor_id") or data.get("user_id") or ""
    notes = data.get("notes", "").strip() or "Session conducted successfully."

    conn = db.get_db_connection()
    intervention = conn.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    if not intervention:
        conn.close()
        return jsonify({"success": False, "message": "Intervention not found."}), 404

    now_iso = datetime.datetime.now().isoformat()
    now_str = datetime.datetime.now().strftime("%b %d, %Y %I:%M %p")

    # Update intervention status
    conn.execute("""
        UPDATE interventions SET
            status = 'Completion Requested',
            completion_requested_by = ?,
            completion_request_notes = ?,
            completion_requested_at = ?
        WHERE id = ?
    """, (mentor_id, notes, now_iso, intervention_id))

    # Fetch mentor and student info
    mentor_user = conn.execute("SELECT display_name FROM users WHERE LOWER(id) = LOWER(?)", (mentor_id,)).fetchone()
    mentor_name = mentor_user["display_name"] if mentor_user else mentor_id
    student = conn.execute("SELECT name FROM students WHERE id = ?", (intervention["student_id"],)).fetchone()
    student_name = student["name"] if student else intervention["student_id"]

    # Dispatch notification to initiator
    initiator_id = intervention["created_by"] or "admin"
    conn.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES (?, ?, 'info', ?, NULL, 0)
    """, (
        f"📝 Completion Review Requested: {intervention['action']}",
        f"{mentor_name} has requested completion approval for {student_name}'s session ({intervention['action']}). Notes: {notes}",
        now_str
    ))

    conn.commit()
    conn.close()
    return jsonify({
        "success": True, 
        "message": f"Completion review request dispatched to session initiator ({initiator_id}) for verification."
    })


@app.route("/api/interventions/<int:intervention_id>/approve-completion", methods=["POST"])
def approve_intervention_completion(intervention_id):
    """Admin or Faculty approves mentor's completion request."""
    data = request.get_json() or {}
    reviewer_id = data.get("reviewer_id") or data.get("user_id") or "admin"

    conn = db.get_db_connection()
    intervention = conn.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    if not intervention:
        conn.close()
        return jsonify({"success": False, "message": "Intervention not found."}), 404

    today_iso = datetime.date.today().isoformat()
    now_iso = datetime.datetime.now().isoformat()
    now_str = datetime.datetime.now().strftime("%b %d, %Y %I:%M %p")

    conn.execute("""
        UPDATE interventions SET
            status = 'Completed',
            completed_date = ?,
            reviewed_by = ?,
            reviewed_at = ?,
            rejection_reason = NULL
        WHERE id = ?
    """, (today_iso, reviewer_id, now_iso, intervention_id))

    # Notify mentor that their completion request was approved
    student = conn.execute("SELECT name FROM students WHERE id = ?", (intervention["student_id"],)).fetchone()
    student_name = student["name"] if student else intervention["student_id"]

    conn.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES (?, ?, 'success', ?, NULL, 0)
    """, (
        f"✅ Session Completion Approved: {intervention['action']}",
        f"Completion verified for {student_name} ({intervention['action']}) by {reviewer_id}. Session is marked as Completed.",
        now_str
    ))

    conn.commit()
    conn.close()
    return jsonify({
        "success": True,
        "message": f"Intervention #{intervention_id} approved and marked as Completed."
    })


@app.route("/api/interventions/<int:intervention_id>/reject-completion", methods=["POST"])
def reject_intervention_completion(intervention_id):
    """Admin or Faculty rejects completion request with a mandatory reason."""
    data = request.get_json() or {}
    reason = (data.get("reason") or "").strip()
    reviewer_id = data.get("reviewer_id") or data.get("user_id") or "admin"

    if not reason:
        return jsonify({"success": False, "message": "A detailed explanation/reason is required when rejecting a completion request."}), 400

    conn = db.get_db_connection()
    intervention = conn.execute("SELECT * FROM interventions WHERE id = ?", (intervention_id,)).fetchone()
    if not intervention:
        conn.close()
        return jsonify({"success": False, "message": "Intervention not found."}), 404

    now_iso = datetime.datetime.now().isoformat()
    now_str = datetime.datetime.now().strftime("%b %d, %Y %I:%M %p")

    conn.execute("""
        UPDATE interventions SET
            status = 'Revision Needed',
            rejection_reason = ?,
            reviewed_by = ?,
            reviewed_at = ?
        WHERE id = ?
    """, (reason, reviewer_id, now_iso, intervention_id))

    student = conn.execute("SELECT name FROM students WHERE id = ?", (intervention["student_id"],)).fetchone()
    student_name = student["name"] if student else intervention["student_id"]

    conn.execute("""
        INSERT INTO notifications (title, message, type, date, student_id, read)
        VALUES (?, ?, 'danger', ?, NULL, 0)
    """, (
        f"⚠️ Session Revision Requested: {intervention['action']}",
        f"Completion review for {student_name} was returned with feedback: '{reason}'. Please review and schedule follow-up.",
        now_str
    ))

    conn.commit()
    conn.close()
    return jsonify({
        "success": True,
        "message": f"Completion request returned for revision. Feedback dispatched to mentor."
    })


@app.route("/api/interventions/enquiries", methods=["GET"])
def get_intervention_enquiries():
    """Returns completion enquiries & review requests for Admin, Faculty, and Mentors."""
    conn = db.get_db_connection()
    user_id = (request.args.get("user_id") or "").lower()
    role = (request.args.get("role") or "").lower()
    status_filter = (request.args.get("status") or "all").lower()
    params = []

    query = """
        SELECT i.*, 
               s.name as student_name, s.course as student_course, s.year as student_year, s.risk as student_risk,
               s.attendance as student_attendance, s.cgpa as student_cgpa,
               u.display_name as mentor_name, u.email as mentor_email,
               creator.display_name as creator_name,
               reviewer.display_name as reviewer_name
        FROM interventions i
        LEFT JOIN students s ON i.student_id = s.id
        LEFT JOIN users u ON LOWER(i.mentor_id) = LOWER(u.id)
        LEFT JOIN users creator ON LOWER(i.created_by) = LOWER(creator.id)
        LEFT JOIN users reviewer ON LOWER(i.reviewed_by) = LOWER(reviewer.id)
        WHERE (i.completion_requested_by IS NOT NULL OR i.status IN ('Completion Requested', 'Revision Needed', 'Completed'))
    """
    if role == "student" and user_id:
        query = """
            SELECT i.*, 
                   s.name as student_name, s.course as student_course, s.year as student_year, s.risk as student_risk,
                   s.attendance as student_attendance, s.cgpa as student_cgpa,
                   u.display_name as mentor_name, u.email as mentor_email,
                   creator.display_name as creator_name,
                   reviewer.display_name as reviewer_name
            FROM interventions i
            LEFT JOIN students s ON i.student_id = s.id
            LEFT JOIN users u ON LOWER(i.mentor_id) = LOWER(u.id)
            LEFT JOIN users creator ON LOWER(i.created_by) = LOWER(creator.id)
            LEFT JOIN users reviewer ON LOWER(i.reviewed_by) = LOWER(reviewer.id)
            WHERE LOWER(i.student_id) = LOWER(?)
        """
        params = [user_id]
        if status_filter == "pending":
            query += " AND i.status IN ('In Progress', 'Pending', 'Completion Requested')"
        elif status_filter == "completed":
            query += " AND i.status = 'Completed'"
    else:
        if status_filter == "pending":
            query += " AND i.status = 'Completion Requested'"
        elif status_filter == "revision":
            query += " AND i.status = 'Revision Needed'"
        elif status_filter == "completed":
            query += " AND i.status = 'Completed' AND i.completion_requested_by IS NOT NULL"

        if role == "mentor" and user_id:
            query += " AND (LOWER(i.mentor_id) = LOWER(?) OR LOWER(i.completion_requested_by) = LOWER(?))"
            params.extend([user_id, user_id])
        elif role == "faculty" and user_id:
            query += " AND (LOWER(i.created_by) = LOWER(?) OR LOWER(i.mentor_id) = LOWER(?))"
            params.extend([user_id, user_id])

    query += " ORDER BY COALESCE(i.completion_requested_at, i.date) DESC, i.id DESC"
    rows = conn.execute(query, tuple(params)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# =====================================================
# 6. AUTONOMOUS AGENT
# =====================================================

@app.route("/api/agent/run-autonomous-loop", methods=["POST"])
def run_autonomous_loop():
    result = agent.run_autonomous_cycle()
    if isinstance(result, dict):
        return jsonify({"success": True, **result})
    return jsonify({
        "success": True,
        "actions_count": len(result),
        "traces": result
    })


@app.route("/api/agent/chat", methods=["POST"])
def agent_chat():
    data = request.get_json() or {}
    user_query = data.get("query", "").strip()
    history = data.get("history", [])
    provider = data.get("provider")

    if not user_query:
        return jsonify({"response": "Please enter a valid query."})

    answer = agent.chat_query(user_query, history=history, provider_override=provider)
    return jsonify({"response": answer})


@app.route("/api/agent/logs", methods=["GET"])
def get_agent_logs():
    conn = db.get_db_connection()
    logs = conn.execute("SELECT * FROM agent_logs ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()
    return jsonify([dict(l) for l in logs])


# =====================================================
# 7. SETTINGS
# =====================================================

def _sync_key_to_env(provider, key):
    """Write submitted API key to .env file and active os.environ."""
    if not key or "•" in key:
        return
    env_map = {
        "gemini": "GEMINI_API_KEY",
        "openai": "OPENAI_API_KEY",
        "groq": "GROQ_API_KEY",
        "openrouter": "OPENROUTER_API_KEY",
        "deepseek": "DEEPSEEK_API_KEY"
    }
    var_name = env_map.get((provider or "").lower())
    if not var_name:
        return

    os.environ[var_name] = key.strip()

    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    try:
        lines = []
        found = False
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

        new_lines = []
        for line in lines:
            if line.strip().startswith(f"{var_name}="):
                new_lines.append(f"{var_name}={key.strip()}\n")
                found = True
            else:
                new_lines.append(line)

        if not found:
            if new_lines and not new_lines[-1].endswith("\n"):
                new_lines[-1] += "\n"
            new_lines.append(f"{var_name}={key.strip()}\n")

        with open(env_path, "w", encoding="utf-8") as f:
            f.writelines(new_lines)
    except Exception as e:
        print(f"[Warning]: Could not write to .env file: {e}")


@app.route("/api/settings", methods=["GET"])
def get_settings():
    """Return settings with API key masked for security."""
    settings = db.get_system_settings()

    # Resolve API key: DB first, then .env fallback per provider
    raw_key = settings.get("api_key", "")
    provider = (settings.get("ai_provider") or "local").lower()
    if not raw_key:
        env_map = {
            "gemini": "GEMINI_API_KEY",
            "openai": "OPENAI_API_KEY",
            "groq": "GROQ_API_KEY",
            "openrouter": "OPENROUTER_API_KEY",
            "deepseek": "DEEPSEEK_API_KEY"
        }
        raw_key = os.environ.get(env_map.get(provider, ""), "")

    # Mask the key: show only last 4 characters
    if raw_key and len(raw_key) > 4:
        settings["api_key"] = "••••••••" + raw_key[-4:]
    elif raw_key:
        settings["api_key"] = "••••"
    else:
        settings["api_key"] = ""

    return jsonify(settings)


def _resolve_api_key(provider):
    """Get the real (unmasked) API key for a given provider from DB or .env."""
    settings = db.get_system_settings()
    raw_key = (settings.get("api_key") or "").strip()
    if not raw_key:
        env_map = {
            "gemini": "GEMINI_API_KEY",
            "openai": "OPENAI_API_KEY",
            "groq": "GROQ_API_KEY",
            "openrouter": "OPENROUTER_API_KEY",
            "deepseek": "DEEPSEEK_API_KEY"
        }
        raw_key = os.environ.get(env_map.get(provider, ""), "")
    return raw_key


@app.route("/api/settings", methods=["POST"])
def update_settings():
    """Save settings to DB and synchronize API keys with .env and os.environ."""
    data = request.get_json() or {}

    provider = (data.get("ai_provider") or "local").lower()

    # If an unmasked API key is submitted, sync it to .env and os.environ
    if "api_key" in data:
        submitted_key = data["api_key"].strip()
        if "•" in submitted_key or not submitted_key:
            del data["api_key"]  # Don't overwrite DB with masked or empty value
        else:
            _sync_key_to_env(provider, submitted_key)

    db.save_system_settings(data)
    return jsonify({"success": True, "message": "Settings updated and synchronized with environment."})


# =====================================================
# 8. NOTIFICATIONS & ANOMALIES
# =====================================================

@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    student_id = request.args.get("student_id")
    include_read = request.args.get("include_read", "true")
    conn = db.get_db_connection()

    if student_id:
        # UAC FIX: Students ONLY see notifications tagged to their own ID
        # No more leaking NULL/global notifications to students
        if include_read == "false":
            rows = conn.execute("SELECT * FROM notifications WHERE student_id = ? AND read = 0 ORDER BY id DESC LIMIT 50", (student_id,)).fetchall()
        else:
            rows = conn.execute("SELECT * FROM notifications WHERE student_id = ? ORDER BY id DESC LIMIT 50", (student_id,)).fetchall()
    else:
        # Admin/Faculty/Mentor see all notifications
        if include_read == "false":
            rows = conn.execute("SELECT * FROM notifications WHERE read = 0 ORDER BY id DESC LIMIT 100").fetchall()
        else:
            rows = conn.execute("SELECT * FROM notifications ORDER BY id DESC LIMIT 100").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/notifications/<int:notif_id>/read", methods=["PUT"])
def mark_notification_read(notif_id):
    """Mark a single notification as read (soft delete)."""
    conn = db.get_db_connection()
    conn.execute("UPDATE notifications SET read = 1 WHERE id = ?", (notif_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Notification #{notif_id} marked as read."})


@app.route("/api/notifications/read-all", methods=["PUT"])
def mark_all_notifications_read():
    """Mark all notifications as read. Optionally filter by student_id."""
    data = request.get_json() or {}
    student_id = data.get("student_id")
    conn = db.get_db_connection()
    if student_id:
        conn.execute("UPDATE notifications SET read = 1 WHERE student_id = ?", (student_id,))
    else:
        conn.execute("UPDATE notifications SET read = 1")
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "All notifications marked as read."})


@app.route("/api/notifications/<int:notif_id>", methods=["DELETE"])
def delete_notification(notif_id):
    """Hard delete a single notification."""
    conn = db.get_db_connection()
    conn.execute("DELETE FROM notifications WHERE id = ?", (notif_id,))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": f"Notification #{notif_id} deleted."})


@app.route("/api/notifications/delete-all", methods=["DELETE"])
def delete_all_notifications():
    """Hard delete all notifications. Optionally filter by student_id."""
    student_id = request.args.get("student_id")
    conn = db.get_db_connection()
    if student_id:
        conn.execute("DELETE FROM notifications WHERE student_id = ?", (student_id,))
    else:
        conn.execute("DELETE FROM notifications")
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "All notifications deleted."})


@app.route("/api/anomalies", methods=["GET"])
def get_anomalies():
    settings = db.get_system_settings()
    attd_threshold = float(settings.get("attendance_threshold", 75))

    conn = db.get_db_connection()
    rows = conn.execute("SELECT * FROM students WHERE attendance < ? OR risk >= 50", (attd_threshold,)).fetchall()

    if not rows:
        conn.close()
        return jsonify([])

    # Batch fetch subject marks for all matching students in ONE query to avoid N+1 bottleneck
    student_ids = [r["id"] for r in rows]
    placeholders = ",".join(["?"] * len(student_ids))
    all_marks = conn.execute(f"SELECT * FROM subject_marks WHERE student_id IN ({placeholders})", student_ids).fetchall()
    conn.close()

    marks_by_student = {}
    for m in all_marks:
        sid = m["student_id"]
        if sid not in marks_by_student:
            marks_by_student[sid] = []
        marks_by_student[sid].append(dict(m))

    anomalies_list = []
    for r in rows:
        st = dict(r)
        marks_list = marks_by_student.get(st["id"], [])
        detected = ai_engine.detect_student_anomalies(st, subject_marks=marks_list)
        if detected:
            anomalies_list.append({
                "student_id": st["id"],
                "student_name": st["name"],
                "attendance": st.get("attendance", 0),
                "cgpa": st.get("cgpa", 0.0),
                "lms_score": st.get("lms_score", st.get("attendance", 0)),
                "course": st.get("course", "CSE"),
                "year": st.get("year", "2nd Year"),
                "risk": st.get("risk", 0),
                "anomalies": detected
            })
    return jsonify(anomalies_list)



# =====================================================
# 9. BULK CSV IMPORT
# =====================================================

@app.route("/api/students/import", methods=["POST"])
def import_students_csv():
    """Bulk import students from CSV. Expects multipart/form-data with 'file' field.
    Required columns: id, name. Optional: gender, course, year, cgpa, credits, attendance, lms_score,
    father, mother, mother_tongue, place, region, country.
    Missing optional fields get sensible defaults."""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded."}), 400

    file = request.files['file']
    filename = file.filename.lower()

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[str(headers[i]).strip()] = cell.value
                    rows.append(row_dict)
            except ImportError:
                return jsonify({"success": False, "message": "Excel support requires openpyxl. Please use CSV format."}), 400
        else:
            return jsonify({"success": False, "message": "Unsupported file format. Use .csv or .xlsx"}), 400

        if not rows:
            return jsonify({"success": False, "message": "File is empty or has no data rows."}), 400

        thresholds = _get_thresholds()
        conn = db.get_db_connection()
        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows):
            row = {k.strip().lower().replace(' ', '_'): (str(v).strip() if v is not None else '') for k, v in row.items()}
            student_id = row.get('id', row.get('student_id', '')).strip()
            name = row.get('name', row.get('full_name', row.get('student_name', ''))).strip()

            if not student_id or not name:
                skipped += 1
                errors.append(f"Row {i+2}: Missing required field 'id' or 'name'")
                continue

            gender = row.get('gender', 'Male') or 'Male'
            course = row.get('course', 'CSE') or 'CSE'
            year = row.get('year', '2nd Year') or '2nd Year'
            try:
                cgpa = float(row.get('cgpa', 7.5) or 7.5)
            except (ValueError, TypeError):
                cgpa = 7.5
            try:
                credits = int(float(row.get('credits', 24) or 24))
            except (ValueError, TypeError):
                credits = 24
            try:
                attendance = int(float(row.get('attendance', 80) or 80))
            except (ValueError, TypeError):
                attendance = 80
            try:
                lms_score = int(float(row.get('lms_score', attendance) or attendance))
            except (ValueError, TypeError):
                lms_score = attendance

            father = row.get('father', row.get("father's_name", 'N/A')) or 'N/A'
            mother = row.get('mother', row.get("mother's_name", 'N/A')) or 'N/A'
            mother_tongue = row.get('mother_tongue', row.get('mothertongue', 'Telugu')) or 'Telugu'
            place = row.get('place', row.get('city', 'Hyderabad')) or 'Hyderabad'
            region = row.get('region', 'South India') or 'South India'
            country = row.get('country', 'India') or 'India'

            risk_info = ai_engine.calculate_risk_score(attendance, cgpa, lms_score, **thresholds)
            risk_score = risk_info["risk_score"]

            try:
                conn.execute("""
                    INSERT OR IGNORE INTO students (id, name, gender, course, year, cgpa, credits, attendance, lms_score, risk, father, mother, mother_tongue, place, region, country)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (student_id, name, gender, course, year, cgpa, credits, attendance, lms_score, risk_score, father, mother, mother_tongue, place, region, country))
                # Create student user account
                email = row.get('email', '').strip() or f"{student_id.lower()}@vignan.ac.in"
                phone = row.get('phone', '').strip()
                conn.execute("INSERT OR IGNORE INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone) VALUES (?,?,?,?,?,?,?,?,?)", (
                    student_id, student_id, "student", name, student_id, None, None, email, phone
                ))
                imported += 1
            except Exception as e:
                skipped += 1
                errors.append(f"Row {i+2} ({student_id}): {str(e)}")

        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(rows),
            "errors": errors[:10],
            "message": f"Imported {imported} students. {skipped} skipped."
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Import failed: {str(e)}"}), 500


@app.route("/api/users/import", methods=["POST"])
def import_users_csv():
    """Bulk import faculty/mentors from CSV. Admin only.
    Required columns: id, role. Optional: display_name, subjects, extra_roles."""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded."}), 400

    file = request.files['file']
    filename = file.filename.lower()

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[str(headers[i]).strip()] = cell.value
                    rows.append(row_dict)
            except ImportError:
                return jsonify({"success": False, "message": "Excel support requires openpyxl. Please use CSV format."}), 400
        else:
            return jsonify({"success": False, "message": "Unsupported file format. Use .csv or .xlsx"}), 400

        if not rows:
            return jsonify({"success": False, "message": "File is empty."}), 400

        conn = db.get_db_connection()
        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows):
            row = {k.strip().lower().replace(' ', '_'): (str(v).strip() if v is not None else '') for k, v in row.items()}
            user_id = row.get('id', row.get('user_id', '')).strip()
            role = row.get('role', 'mentor').strip().lower()

            if not user_id:
                skipped += 1
                errors.append(f"Row {i+2}: Missing 'id'")
                continue

            if role not in ('faculty', 'mentor', 'admin'):
                role = 'mentor'

            display_name = row.get('display_name', row.get('name', user_id)) or user_id
            password = row.get('password', user_id) or user_id
            subjects = row.get('subjects', row.get('assigned_subjects', '')) or ''
            extra_roles = row.get('extra_roles', row.get('responsibilities', '')) or ''
            email = row.get('email', '').strip()
            phone = row.get('phone', '').strip()

            try:
                conn.execute("INSERT OR IGNORE INTO users (id, password, role, display_name, linked_student_id, subjects, extra_roles, email, phone) VALUES (?,?,?,?,?,?,?,?,?)", (
                    user_id, password, role, display_name, None, subjects, extra_roles, email, phone
                ))
                imported += 1
            except Exception as e:
                skipped += 1
                errors.append(f"Row {i+2} ({user_id}): {str(e)}")

        conn.commit()
        conn.close()
        return jsonify({
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(rows),
            "errors": errors[:10],
            "message": f"Imported {imported} users. {skipped} skipped."
        })
    except Exception as e:
        return jsonify({"success": False, "message": f"Import failed: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
